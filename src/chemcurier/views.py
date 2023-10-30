from django.shortcuts import render
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect
from . import models
from . import forms
from prices import models as prices_models
from prices import forms as prices_forms
from homepage.templatetags import my_tags as homepage_my_tags
from django.views.generic import DetailView, View
import datetime
import pandas as pd
import openpyxl 
import os




class ChemcourierTableModelDetailView(DetailView):
    model = models.ChemCourierTableModel
    template_name = 'chemcurier/chemcurier.html'
    
    def get_object(self, queryset=None):
        chemcourier_table = models.ChemCourierTableModel.objects.get_or_create(chemcourier_table='chemcourier_table')[0]
        return chemcourier_table
    
    def get_context_data(self, **kwargs):       
        context = super().get_context_data(**kwargs)
        obj = context.get('object')  

        context['chemcourier_table'] = obj
        # 1. 
        # 1.1 получить период дат (месяц-год), доступных в спарсенной базе Хим Курьер
        # сфомированно в форме

        #1.2 форма для ввода периорда + получение месяца и года для поиска в базе Chemcurier объектов:
        chosen_month_num = 1
        chosen_year_num = 1
        if forms.INITIAL_PERIOD:                                                            # если пользователь выбрал период
            period_form = forms.PeriodForm(initial={'periods': forms.INITIAL_PERIOD})
            chosen_period = datetime.datetime.strptime(forms.INITIAL_PERIOD, '%Y-%m-%d').date()         
            chosen_year_num = chosen_period.year
            chosen_month_num = chosen_period.month   
        else: 
            try:                                                                               # если пользователь не выбрал период
                period_form = forms.PeriodForm()
                last_period = forms.PERIODS[-1] 
                chosen_period = last_period[0]
                chosen_year_num = chosen_period.year
                chosen_month_num = chosen_period.month  
            except:
                period_form = forms.PeriodForm()
        context['period_form'] = period_form

        #1.3 для заголовка отображение выбранного периода в форме месяц - год
        NUMBER_TO_MONTH_DIC_FOR_HEADER = { 1 : 'январь', 2: 'февраль', 3 : 'март', 4 : 'апрель', 5 : 'май', 6 :'июнь', 7 : 'июль', 8 : 'август', 9 : 'сентябрь', 10 : 'октябрь', 11 : 'ноябрь', 12 : 'декабрь'}
        period_is = NUMBER_TO_MONTH_DIC_FOR_HEADER.get(chosen_month_num) + ' ' + str(chosen_year_num) 
        context['period_is'] = period_is

        # 2 получение типоразмера для отбора (создание формы):
        tyresizes_form = forms.TyreSizeForm()
        tyresize_to_check = None
        if forms.INITIAL_TYREISIZE:
    #        print('1===', forms.INITIAL_TYREISIZE )
            tyresizes_form = forms.TyreSizeForm(initial={'tyresizes': forms.INITIAL_TYREISIZE})
            tyresize_to_check = forms.INITIAL_TYREISIZE
        else:
    #        print('2===')
            if forms.TYRESIZES:
                tyresizes_form = forms.TyreSizeForm()
                tyresize_to_check = forms.TYRESIZES[0]      #берем первый из списка    
        context['tyresizes_form'] = tyresizes_form
        context['current_tyresize'] = tyresize_to_check
        # 3 (необязательные)
        # 3.1 получение бренда для отбора (создание формы):
        tyrebrands_to_check = None
        if forms.INITIAL_BRANDS and forms.INITIAL_BRANDS != '-':
            tyrebrands_form = forms.BrandForm(initial={'tyrebrands': forms.INITIAL_BRANDS})
            tyrebrands_to_check = forms.INITIAL_BRANDS
        else:    
            tyrebrands_form = forms.BrandForm()
        context['tyrebrands_form'] = tyrebrands_form
        # 3.2 получение получателя для отбора (создание формы):
        recievers_to_check = None
        if forms.INITIAL_RECIEVER and forms.INITIAL_RECIEVER != '-':
            recievers_form = forms.RecieverForm(initial={'recievers': forms.INITIAL_RECIEVER})
            recievers_to_check = forms.INITIAL_RECIEVER
        else:    
            recievers_form = forms.RecieverForm()
        context['recievers_form'] = recievers_form      
        # 3.3 получение страну производства для отбора (создание формы):
        prod_countrys_to_check = None
        if forms.INITIAL_PRODCOUTRYS and forms.INITIAL_PRODCOUTRYS != '-':
            prod_countrys_form = forms.ProdCountryForm(initial={'prod_countrys': forms.INITIAL_PRODCOUTRYS})
            prod_countrys_to_check = forms.INITIAL_PRODCOUTRYS
        else:    
            prod_countrys_form = forms.ProdCountryForm()
        context['prod_countrys_form'] = prod_countrys_form   

        # 3.4 получение группу для отбора (создание формы):
        prod_groups_to_check = None
        if forms.INITIAL_GROUPS and forms.INITIAL_GROUPS != '-':
            groups_form = forms.GroupForm(initial={'prod_groups': forms.INITIAL_GROUPS})
            prod_groups_to_check = forms.INITIAL_GROUPS
        else:    
            groups_form = forms.GroupForm()
        context['groups_form'] = groups_form   

        # 4 ДЛЯ ПОЛУЧЕНИЯ ВАЛЮТЫ ПО КУРСУ НБ РБ НА ДАТУ    
        #print('prices_models.CURRENCY_DATE_GOT_FROM_USER ===========', prices_models.CURRENCY_DATE_GOT_FROM_USER, 'prices_models.CURRENCY_IS_CLEANED ====', prices_models.CURRENCY_IS_CLEANED)   
        curr_value = None
        currency = None
        shown_date = None
        if prices_models.CURRENCY_ON_DATE is False:                         # запускаем получение курса валют с НБ РБ только раз за день
            try:
                cu, cu_val, sh_date = homepage_my_tags.currency_on_date()
                if datetime.datetime.today().strftime("%Y-%m-%d") == sh_date:       # если на сегодняшнюю дату результат получен - то более не запрашивать
                    currency, curr_value, shown_date = cu, cu_val, sh_date
                    prices_models.CURRENCY_IS_CLEANED = currency, curr_value, shown_date   # записываем полученные значения
                    prices_models.CURRENCY_ON_DATE = True
            except:
                currency, curr_value, shown_date = homepage_my_tags.currency_on_date()     # если что -то пошло не так - берем данные с сайта  
                prices_models.CURRENCY_IS_CLEANED = currency, curr_value, shown_date
                prices_models.CURRENCY_ON_DATE = True  
        if prices_models.CURRENCY_DATE_GOT_FROM_USER:                              # если пользователь вводит данные (получить курс на определенную дату): #
            if prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED:                  # если только что получал данные на эту дату - то не надо запускать фунцкию - взять что уже собрано
                try:
                    currency_already, curr_value_already, shown_date_already = prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED
                    if shown_date_already == prices_models.CURRENCY_DATE_GOT_FROM_USER:
                       currency_already, curr_value_already, shown_date_already = prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED 
                    else:
                        currency, curr_value, shown_date = homepage_my_tags.currency_on_date()
                        prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED = currency, curr_value, shown_date
                except:
                    currency, curr_value, shown_date = homepage_my_tags.currency_on_date()
                    prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED = currency, curr_value, shown_date
                if shown_date_already == prices_models.CURRENCY_DATE_GOT_FROM_USER:
                   currency, curr_value, shown_date = currency_already, curr_value_already, shown_date_already
            else:                                                           # если ничего - тогда обращаемся к функции:
                currency, curr_value, shown_date = homepage_my_tags.currency_on_date()
                prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED = currency, curr_value, shown_date

        #prices_models.CURRENCY_VALUE_RUB = curr_value / 100
        if curr_value:
            prices_models.CURRENCY_VALUE_USD = curr_value 
        else:
            prices_models.CURRENCY_VALUE_USD = 0

        currency_input_form = prices_forms.CurrencyDateInputForm()
        context['currency_input_form'] = currency_input_form
        context['currency'] = currency
        context['curr_value'] = curr_value
        date_exist_true = None
        if shown_date:
            date_exist_true = datetime.datetime.strptime(shown_date, "%Y-%m-%d").date()
        else:
            date_exist_true = datetime.date.today()
        currency_input_form = prices_forms.CurrencyDateInputForm()       
        currency_input_form.fields['chosen_date_for_currency'].initial = date_exist_true                        # !!! ЭТО БАЗА
        context['currency_input_form'] = currency_input_form
        ## END ДЛЯ ПОЛУЧЕНИЯ ВАЛЮТЫ ПО КУРСУ НБ РБ НА ДАТУ

        # 5 получить все объекты химкурьер за исключением нулевых значений (шт. деньги) на дату:
        get_chem_courier_objects_from_base = prices_models.ChemCurierTyresModel.objects.filter(data_month_chem__month=chosen_month_num, data_month_chem__year=chosen_year_num, tyre_size_chem=tyresize_to_check).exclude(average_price_in_usd__isnull=True)      
        # 5.1 доп. проверки:
        if tyrebrands_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(producer_chem=tyrebrands_to_check)
        if recievers_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(reciever_chem=recievers_to_check)
        if prod_countrys_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(prod_country=prod_countrys_to_check)
        if prod_groups_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(group_chem__tyre_group=prod_groups_to_check)

        models.CHEM_PNJ_IN_TABLE_LIST = get_chem_courier_objects_from_base
        context['get_chem_courier_objects_from_base'] = get_chem_courier_objects_from_base


        ####### 6 СКАЧИВАНИЕ ФАЙЛА - EXCEL ТАБЛИЦА с данными из таблицы
        if forms.CHEMCOURIER_EXCEL_CREATE is True:
            if type(tyresize_to_check) is tuple:
                try:
                    tyresize_to_check_name = str(tyresize_to_check).replace('/', ' ')
                except:
                    tyresize_to_check_name = ' '
            else:
                tyresize_to_check_name = tyresize_to_check.replace('/', ' ') # убрать / из названия для excel вкладки    

            from openpyxl import Workbook
            wb = Workbook()
            if wb['Sheet']:                                         # если есть вкладка с именем 'Sheet' - поменнять на название типоразмера
                sheet_named_sheet = wb['Sheet']
                sheet_named_sheet.title = tyresize_to_check_name
                excel_sheet = sheet_named_sheet
            else:                                                   # если нет  вкладки с именеми 'Sheet' - созать с названием типоразмера
                wb.create_sheet(tyresize_to_check_name)
                excel_sheet = wb[tyresize_to_check_name]

            ###### получить количество объектов:
            row_value = 0
            for ob_val in get_chem_courier_objects_from_base:
                row_value += 1

            #генератор для хождения по строкам:    
            def raw_generator(n, stop):
                while True:
                    if n > stop:
                        raise StopIteration
                    yield n
                    n += 1
            max_raw = row_value + 2  # (+1 строка для ИТОГО и записи что согласно хим курьеру)
            i = 3       # с 3  строки        
            row_curr = raw_generator(i, max_raw+i)

            last_row_in_table_num = 0

            excel_sheet['A1'] = tyresize_to_check_name
            for chem_obj in get_chem_courier_objects_from_base:

                    val = next(row_curr)
                    excel_sheet['A2'] = 'Бренд'
                    excel_sheet.cell(row=val, column=1).value = chem_obj.producer_chem
                    excel_sheet.cell(row=val+1, column=1).value = 'ИТОГО'

                    excel_sheet['B2'] = 'Получатель'
                    excel_sheet.cell(row=val, column=2).value = chem_obj.reciever_chem

                    excel_sheet['C2'] = 'Страна производства'
                    excel_sheet.cell(row=val, column=3).value = chem_obj.prod_country

                    #excel_sheet['G1'] = 'Param'
                    #str_obj_param_list = []
                    #str_obj_param_str = ''
                    #for type_obj in sales_obj.tyre.tyre_type.all():
                    #    str_obj_param_list.append(type_obj.tyre_type)
                    #str_obj_param_str = ', '.join(str_obj_param_list)
                    ##print(str_obj_param_str)
                    #excel_sheet.cell(row=val, column=7).value = str_obj_param_str

                    excel_sheet['D2'] = 'Группа шин'
                    excel_sheet.cell(row=val, column=4).value = chem_obj.group_chem.tyre_group

                    excel_sheet['E2'] = 'Объем поставки, шт.'
                    excel_sheet.cell(row=val, column=5).value = chem_obj.obj_val_on_moth_chem_for_table()
                    excel_sheet.cell(row=val+1, column=5).value = obj.num_summ()[0]

                    excel_sheet['F2'] = 'Объем поставки, USD'
                    excel_sheet.cell(row=val, column=6).value = chem_obj.obj_money_on_moth_chem_for_table()
                    excel_sheet.cell(row=val+1, column=6).value = obj.num_summ()[1]

                    excel_sheet['G2'] = 'Средневзвешенная цена реализации единицы товара, USD'
                    excel_sheet.cell(row=val, column=7).value = chem_obj.obj_average_price_in_usd_chem_for_table()
                    excel_sheet.cell(row=val+1, column=7).value = obj.num_summ()[2]

                    excel_sheet['H2'] = 'Средневзвешенная цена реализации единицы товара, бел.руб.'
                    excel_sheet.cell(row=val, column=8).value = chem_obj.average_from_usd_to_bel()
                    excel_sheet.cell(row=val+1, column=8).value = obj.num_summ()[3]

                    last_row_in_table_num = val+1
            excel_sheet.cell(row=last_row_in_table_num+2, column=1).value = 'Согласно данным информационно-аналитическго агентства «Хим-Курьер»'   

            #  Для отрисовки лийний у таблицы:
            from openpyxl.styles.borders import Border, Side
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
    #        excel_sheet.cell(row=3, column=2).border = thin_border
            #for i in excel_sheet['A1':'D8']:
            cell_range = excel_sheet['A2':f'H{last_row_in_table_num}']
            for row in cell_range:
                for cell in row:
                    cell.border = thin_border

            #wb.save("Chemcourier.xlsx")
            #wb.save('/home/user/Desktop/FileName.xlsx')

            # сохранение в папку DOWNLOADS пользователя:
            user_path_is = os.path.expanduser("~")
            #print('path_is', user_path_is)
            download_folder_exist = user_path_is + "/Downloads/"                            # создать файл в папке Downloads
            if os.path.exists(download_folder_exist):
                users_download_path = user_path_is + "/Downloads/" + 'Chemcourier.xlsx'
                dowload_to = os.path.normpath(users_download_path)     
            else:                                                                           # если папки Downloads нет - создать файл в корневой папке пользователя
                users_download_path = user_path_is  + 'Chemcourier.xlsx'
                dowload_to = os.path.normpath(users_download_path)


            def not_in_use(filename):                                               # проверка - используется ли в данный момент файл (если он откурыт -то будет попытка его переписыть с ошибкой - надо избеать)
                    try:
                        os.rename(filename,filename)
                        return True
                    except:  
                        return False     
            if os.path.exists(dowload_to):          
                if not_in_use(dowload_to):
                    wb.save(dowload_to)
            else:
                wb.save(dowload_to)

            wb.close()      
    

        ####### END СКАЧИВАНИЕ ФАЙЛА - EXCEL ТАБЛИЦА с данными из таблицы

        return context  
    
class ChemcourierTableModelUpdateView(View):

    def post(self, request):

    #    print(request.POST, 'TTTH')
        # 1. получаем период от пользователя для поиска и установки initial значения в дате формы выбора
        get_period = request.POST.get('periods') 
        if get_period:
            forms.INITIAL_PERIOD = get_period
        # 2. получаем типоразмер от пользователя для поиска и установки initial значения в типоразмере формы выбора    
        get_tyresize = request.POST.get('tyresizes') 
        if get_tyresize:
            forms.INITIAL_TYREISIZE = get_tyresize

        # 3. получаем бренд от пользователя для поиска и установки initial значения в бренде формы выбора    
        get_tyrebrand = request.POST.get('tyrebrands') 
        if get_tyrebrand and get_tyrebrand != '-':
            forms.INITIAL_BRANDS = get_tyrebrand
        else:
            forms.INITIAL_BRANDS = None

        # 4. получаем получателя от пользователя для поиска и установки initial значения в получателях формы выбора    
        get_reciever = request.POST.get('recievers') 
        if get_reciever and get_reciever != '-':
            forms.INITIAL_RECIEVER = get_reciever
        else:
            forms.INITIAL_RECIEVER = None

        # 5. получаем страну производства от пользователя для поиска и установки initial значения в странах производства формы выбора    
        get_prod_countrys = request.POST.get('prod_countrys') 
        if get_prod_countrys and get_prod_countrys != '-':
            forms.INITIAL_PRODCOUTRYS = get_prod_countrys
        else:
            forms.INITIAL_PRODCOUTRYS = None

        # 6. получаем группу производства от пользователя для поиска и установки initial значения в группах формы выбора    
        get_prod_groups = request.POST.get('prod_groups') 
        if get_prod_groups and get_prod_groups != '-':
            forms.INITIAL_GROUPS = get_prod_groups
        else:
            forms.INITIAL_GROUPS = None

        # 7. курс валют на дату
        chosen_date_for_currency_year = request.POST.getlist('chosen_date_for_currency_year') 
        chosen_date_for_currency_month = request.POST.getlist('chosen_date_for_currency_month') 
        chosen_date_for_currency_day = request.POST.getlist('chosen_date_for_currency_day') 
        chosen_date_for_currency = chosen_date_for_currency_year + chosen_date_for_currency_month + chosen_date_for_currency_day
        if chosen_date_for_currency:
            chosen_date_for_currency = '-'.join(str(x) for x in chosen_date_for_currency)
            check_date = datetime.datetime.strptime(chosen_date_for_currency, "%Y-%m-%d").date()        #  если пользователем введена дана превышающая текущую для получения курса валют то нао скинуть на сегодня:
            if check_date > datetime.datetime.now().date():
                prices_models.CURRENCY_ON_DATE is False
            else:
                prices_models.CURRENCY_DATE_GOT_FROM_USER = chosen_date_for_currency
                prices_models.CURRENCY_ON_DATE is True

        # 8. создать EXCEL CHEMCOURIER
        chemcuorier_download = request.POST.getlist('chemcuorier_download') 
        if chemcuorier_download:
            forms.CHEMCOURIER_EXCEL_CREATE = True


        return HttpResponseRedirect(reverse_lazy('chemcurier:chemcurier_table'))


class ChemcourierProgressiveTableModelDetailView(DetailView):
    model = models.ChemCourierTableModel
    template_name = 'chemcurier/chemcurier_progressive.html'
    
    def get_object(self, queryset=None):
        chemcourier_table = models.ChemCourierTableModel.objects.get_or_create(chemcourier_table='chemcourier_table')[0]
        return chemcourier_table
    
    def get_context_data(self, **kwargs):       
        context = super().get_context_data(**kwargs)
        obj = context.get('object')  

        context['chemcourier_table'] = obj
        # 1. 
        # 1.1 получить период дат (месяц-год), доступных в спарсенной базе Хим Курьер
        # сфомированно в форме

        #1.2 форма для ввода периорда + получение месяца и года для поиска в базе Chemcurier объектов:
        chosen_per_num_start = None       
        chosen_per_num_end = None 
        if forms.INITIAL_PERIOD_START or forms.INITIAL_PERIOD_END:
            if forms.INITIAL_PERIOD_START:                                                            # если пользователь выбрал период
                period_form_start = forms.StartPeriodForm(initial={'periods_start': forms.INITIAL_PERIOD_START})
                chosen_period = datetime.datetime.strptime(forms.INITIAL_PERIOD_START, '%Y-%m-%d').date()         
                chosen_per_num_start = chosen_period
            if forms.INITIAL_PERIOD_END:      
                period_form_end = forms.EndPeriodForm(initial={'periods_end': forms.INITIAL_PERIOD_END}) 
                chosen_period = datetime.datetime.strptime(forms.INITIAL_PERIOD_END, '%Y-%m-%d').date()         
                chosen_per_num_end = chosen_period     
        else: 
            try:                                                                               # если пользователь не выбрал период
                period_form_start = forms.StartPeriodForm()
                period_form_end = forms.EndPeriodForm()
                last_period = forms.PERIODS[-1] 
                chosen_period = last_period[0]

                chosen_per_num_start = chosen_period 
                chosen_per_num_end = chosen_per_num_start        
            except:
                period_form_start = forms.StartPeriodForm()
                period_form_end = forms.EndPeriodForm()
 
        context['period_form_start'] = period_form_start
        context['period_form_end'] = period_form_end

        # 2 получение типоразмера для отбора (создание формы):
        tyresizes_form = forms.TyreSizeForm()
        tyresize_to_check = None
        if forms.INITIAL_TYREISIZE:
    #        print('1===', forms.INITIAL_TYREISIZE )
            tyresizes_form = forms.TyreSizeForm(initial={'tyresizes': forms.INITIAL_TYREISIZE})
            tyresize_to_check = forms.INITIAL_TYREISIZE
        else:
    #        print('2===')
            if forms.TYRESIZES:
                tyresizes_form = forms.TyreSizeForm()
                tyresize_to_check = forms.TYRESIZES[0]      #берем первый из списка    
        context['tyresizes_form'] = tyresizes_form
        context['current_tyresize'] = tyresize_to_check
        # 3 (необязательные)
        # 3.1 получение бренда для отбора (создание формы):
        tyrebrands_to_check = None
        if forms.INITIAL_BRANDS and forms.INITIAL_BRANDS != '-':
            tyrebrands_form = forms.BrandForm(initial={'tyrebrands': forms.INITIAL_BRANDS})
            tyrebrands_to_check = forms.INITIAL_BRANDS
        else:    
            tyrebrands_form = forms.BrandForm()
        context['tyrebrands_form'] = tyrebrands_form
        # 3.2 получение получателя для отбора (создание формы):
        recievers_to_check = None
        if forms.INITIAL_RECIEVER and forms.INITIAL_RECIEVER != '-':
            recievers_form = forms.RecieverForm(initial={'recievers': forms.INITIAL_RECIEVER})
            recievers_to_check = forms.INITIAL_RECIEVER
        else:    
            recievers_form = forms.RecieverForm()
        context['recievers_form'] = recievers_form      
        # 3.3 получение страну производства для отбора (создание формы):
        prod_countrys_to_check = None
        if forms.INITIAL_PRODCOUTRYS and forms.INITIAL_PRODCOUTRYS != '-':
            prod_countrys_form = forms.ProdCountryForm(initial={'prod_countrys': forms.INITIAL_PRODCOUTRYS})
            prod_countrys_to_check = forms.INITIAL_PRODCOUTRYS
        else:    
            prod_countrys_form = forms.ProdCountryForm()
        context['prod_countrys_form'] = prod_countrys_form   

        # 3.4 получение группу для отбора (создание формы):
        prod_groups_to_check = None
        if forms.INITIAL_GROUPS and forms.INITIAL_GROUPS != '-':
            groups_form = forms.GroupForm(initial={'prod_groups': forms.INITIAL_GROUPS})
            prod_groups_to_check = forms.INITIAL_GROUPS
        else:    
            groups_form = forms.GroupForm()
        context['groups_form'] = groups_form   

        # 4 ДЛЯ ПОЛУЧЕНИЯ ВАЛЮТЫ ПО КУРСУ НБ РБ НА ДАТУ    
        #print('prices_models.CURRENCY_DATE_GOT_FROM_USER ===========', prices_models.CURRENCY_DATE_GOT_FROM_USER, 'prices_models.CURRENCY_IS_CLEANED ====', prices_models.CURRENCY_IS_CLEANED)   
        curr_value = None
        currency = None
        shown_date = None
        if prices_models.CURRENCY_ON_DATE is False:                         # запускаем получение курса валют с НБ РБ только раз за день
            try:
                cu, cu_val, sh_date = homepage_my_tags.currency_on_date()
                if datetime.datetime.today().strftime("%Y-%m-%d") == sh_date:       # если на сегодняшнюю дату результат получен - то более не запрашивать
                    currency, curr_value, shown_date = cu, cu_val, sh_date
                    prices_models.CURRENCY_IS_CLEANED = currency, curr_value, shown_date   # записываем полученные значения
                    prices_models.CURRENCY_ON_DATE = True
            except:
                currency, curr_value, shown_date = homepage_my_tags.currency_on_date()     # если что -то пошло не так - берем данные с сайта  
                prices_models.CURRENCY_IS_CLEANED = currency, curr_value, shown_date
                prices_models.CURRENCY_ON_DATE = True  
        if prices_models.CURRENCY_DATE_GOT_FROM_USER:                              # если пользователь вводит данные (получить курс на определенную дату): #
            if prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED:                  # если только что получал данные на эту дату - то не надо запускать фунцкию - взять что уже собрано
                try:
                    currency_already, curr_value_already, shown_date_already = prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED
                    if shown_date_already == prices_models.CURRENCY_DATE_GOT_FROM_USER:
                       currency_already, curr_value_already, shown_date_already = prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED 
                    else:
                        currency, curr_value, shown_date = homepage_my_tags.currency_on_date()
                        prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED = currency, curr_value, shown_date
                except:
                    currency, curr_value, shown_date = homepage_my_tags.currency_on_date()
                    prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED = currency, curr_value, shown_date
                if shown_date_already == prices_models.CURRENCY_DATE_GOT_FROM_USER:
                   currency, curr_value, shown_date = currency_already, curr_value_already, shown_date_already
            else:                                                           # если ничего - тогда обращаемся к функции:
                currency, curr_value, shown_date = homepage_my_tags.currency_on_date()
                prices_models.CURRENCY_DATE_GOT_FROM_USER_CLEANED = currency, curr_value, shown_date

        #prices_models.CURRENCY_VALUE_RUB = curr_value / 100
        if curr_value:
            prices_models.CURRENCY_VALUE_USD = curr_value 
        else:
            prices_models.CURRENCY_VALUE_USD = 0

        currency_input_form = prices_forms.CurrencyDateInputForm()
        context['currency_input_form'] = currency_input_form
        context['currency'] = currency
        context['curr_value'] = curr_value
        date_exist_true = None
        if shown_date:
            date_exist_true = datetime.datetime.strptime(shown_date, "%Y-%m-%d").date()
        else:
            date_exist_true = datetime.date.today()
        currency_input_form = prices_forms.CurrencyDateInputForm()       
        currency_input_form.fields['chosen_date_for_currency'].initial = date_exist_true                        # !!! ЭТО БАЗА
        context['currency_input_form'] = currency_input_form
        ## END ДЛЯ ПОЛУЧЕНИЯ ВАЛЮТЫ ПО КУРСУ НБ РБ НА ДАТУ


        # 5 получить все объекты химкурьер за исключением нулевых значений (шт. деньги) на дату:
        get_chem_courier_objects_from_base = prices_models.ChemCurierTyresModel.objects.filter(data_month_chem__range=(chosen_per_num_start, chosen_per_num_end), tyre_size_chem=tyresize_to_check).exclude(average_price_in_usd__isnull=True)      
     
        # 5.1 доп. проверки:
        if tyrebrands_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(producer_chem=tyrebrands_to_check)
        if recievers_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(reciever_chem=recievers_to_check)
        if prod_countrys_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(prod_country=prod_countrys_to_check)
        if prod_groups_to_check:   
            get_chem_courier_objects_from_base = get_chem_courier_objects_from_base.filter(group_chem__tyre_group=prod_groups_to_check)

        models.CHEM_PNJ_IN_TABLE_LIST = get_chem_courier_objects_from_base
    #    context['get_chem_courier_objects_from_base'] = get_chem_courier_objects_from_base
        
        # 6 в отрисовку таблицы 
        context['get_chem_courier_objects_from_base'] = obj.table_content_creation()[0]
        #print('!!!!!!!!', context['get_chem_courier_objects_from_base'])

        # 7 количество столбцов именно сданными на дату:
        headers_len = []
        for k, v in context['get_chem_courier_objects_from_base'].items():
            for keyy in v.keys():
                headers_len.append(keyy)
            break
        headers_len = sorted(headers_len)       # даты заголовков по порядку
        #print('headers_len', headers_len)
        NUMBER_TO_MONTH_DIC_FOR_HEADER = { 1 : 'январь', 2: 'февраль', 3 : 'март', 4 : 'апрель', 5 : 'май', 6 :'июнь', 7 : 'июль', 8 : 'август', 9 : 'сентябрь', 10 : 'октябрь', 11 : 'ноябрь', 12 : 'декабрь'}
        headers_len_dict = {}
        for dattta in headers_len:
            month_is = dattta.month
            year_is = dattta.year
            headers_len_dict[dattta] = NUMBER_TO_MONTH_DIC_FOR_HEADER.get(month_is) + ' ' + str(year_is) 

        context['headers_len'] = headers_len_dict

        # 7 для заголовка таблицы отображение рассматриваемого периода
        try:
            start_ddate_period_is = list(headers_len_dict.keys())[0]
            start_ddate_period_is_month_is = start_ddate_period_is.month
            start_ddate_period_is_year_is = start_ddate_period_is.year
            end_ddate_period_is= list(headers_len_dict.keys())[-1]
            end_ddate_period_is_month_is = end_ddate_period_is.month
            end_ddate_period_is_year_is = end_ddate_period_is.year  
            context['period_start'] = NUMBER_TO_MONTH_DIC_FOR_HEADER.get(start_ddate_period_is_month_is) + ' ' + str(start_ddate_period_is_year_is) 
            context['period_end'] = NUMBER_TO_MONTH_DIC_FOR_HEADER.get(end_ddate_period_is_month_is) + ' ' + str(end_ddate_period_is_year_is) 
        except:
            context['period_start'] = 'начало периода'
            context['period_end'] = 'конец периода'
          
        # 8 ИТОГО     
        context['itogo'] = obj.table_content_creation()[1]

        # 9 ИТОГО по правым колонкам (за весь период)
        def num_summ_right_columns_result():
            num_summ_right_columns_result = {}
            for key, val in context['get_chem_courier_objects_from_base'].items():
                item_val = 0
                sum_val = 0
                aver_val = 0
                for v in val.values():
                    if v[0] != ' ':
                        item_val += int(v[0].replace(' ', ''))
                        sum_val += float(v[1].replace(' ', ''))
                        aver_val = sum_val / item_val
                item_val = '{0:,}'.format(item_val).replace(',', ' ')
                sum_val = '{0:,}'.format(sum_val).replace(',', ' ')
                aver_val = float('{:.2f}'.format(aver_val))
                aver_val = '{0:,}'.format(aver_val).replace(',', ' ')
                num_summ_right_columns_result[key] = item_val, sum_val, aver_val
            return num_summ_right_columns_result 
        context['num_summ_right_columns_result'] = num_summ_right_columns_result()
        #print('context[num_summ_right_columns_result]', context['num_summ_right_columns_result'])

        # 9.1 ИТОГО по правым колонкам (за весь период) СУММА СУММ (правая ниюняя самая итоговая в таблице)
        values_in_togo_list = []
        for val in context['itogo'].values():
            values_in_togo_list.append(val)
        num_summ_right_columns_result_itogo = []
        item_val = 0
        sum_val = 0
        aver_val = 0
        for val in values_in_togo_list:
            item_val1, item_val2, item_val3 = val
            item_val += int(item_val1.replace(' ', ''))
            sum_val += float(item_val2.replace(' ', ''))
            aver_val = sum_val / item_val
        item_val = '{0:,}'.format(item_val).replace(',', ' ')
        sum_val = float('{:.2f}'.format(sum_val))
        sum_val = '{0:,}'.format(sum_val).replace(',', ' ')
        aver_val = float('{:.2f}'.format(aver_val))
        aver_val = '{0:,}'.format(aver_val).replace(',', ' ')    
        num_summ_right_columns_result_itogo = item_val, sum_val, aver_val
        context['num_summ_right_columns_result_itog_sum'] = num_summ_right_columns_result_itogo
        #print('lost_inside_combojia', context['num_summ_right_columns_result_itog_sum'])
    

        ####### 10 СКАЧИВАНИЕ ФАЙЛА - EXCEL ТАБЛИЦА с данными из таблицы
        if forms.CHEMCOURIER_PROGRESSIVE_EXCEL_CREATE is True:
            if type(tyresize_to_check) is tuple:
                try:
                    tyresize_to_check_name = str(tyresize_to_check).replace('/', ' ')
                except:
                    tyresize_to_check_name = ' '
            else:
                tyresize_to_check_name = tyresize_to_check.replace('/', ' ') # убрать / из названия для excel вкладки  

            #from openpyxl.utils.cell import coordinate_from_string, column_index_from_string      # для координат из А в 1 !!!!!!
            from openpyxl.utils import get_column_letter                                           # для координат из 1 в А !!!!!!
            from openpyxl import Workbook
            wb = Workbook()
            if wb['Sheet']:                                         # если есть вкладка с именем 'Sheet' - поменнять на название типоразмера
                sheet_named_sheet = wb['Sheet']
                sheet_named_sheet.title = tyresize_to_check_name
                excel_sheet = sheet_named_sheet
            else:                                                   # если нет  вкладки с именеми 'Sheet' - созать с названием типоразмера
                wb.create_sheet(tyresize_to_check_name)
                excel_sheet = wb[tyresize_to_check_name]
            ###### получить количество объектов:
            row_value = 0
            for ob_val in get_chem_courier_objects_from_base:
                row_value += 1
            #генератор для хождения по строкам:    
            def raw_generator(n, stop):
                while True:
                    if n > stop:
                        raise StopIteration
                    yield n
                    n += 1
            max_raw = row_value + 2  # (+1 строка для ИТОГО и записи что согласно хим курьеру)
            i = 3       # с 3  строки        
            row_curr = raw_generator(i, max_raw+i)
            last_row_in_table_num = 0

            headers_len_periods_num = len(headers_len) * 3      # количество столбцов с данными по периодам

            lolumns_list = []
            columns_dict = {} 
            count_to_three = 0
            periods_count = 0
            all_columns_names = []
            ########### ОЧЕНЬ КЛАСЧНЫЙ СПОСОБ ХОДИТЬ ПО ЯЧЕЙКАМ
            counter = 0 
            head_per_count_num = 4 + headers_len_periods_num                                                        
            for column in range(4,head_per_count_num):       
                column_letter = get_column_letter(column)
    #            for row in range(1,11):
    #                counter = counter +1
    #                excel_sheet[column_letter + str(row)] = counter       
                    #print('**', column_letter + str(row))
    #            print('**', column_letter )
                count_to_three += 1
                lolumns_list.append(column_letter)
                all_columns_names.append(column_letter)
                if count_to_three == 3:
                    columns_dict[periods_count] = lolumns_list
                    count_to_three = 0
                    periods_count += 1
                    lolumns_list = []
#            print('=======', columns_dict)
            ########### END ОЧЕНЬ КЛАСЧНЫЙ СПОСОБ ХОДИТЬ ПО ЯЧЕЙКАМ


            all_columns_names_reversed = list(reversed(all_columns_names))
            num_to_abc_convert_dict = {}
            for column_num in range(4,head_per_count_num): 
                num_to_abc_convert_dict[column_num] = all_columns_names_reversed.pop()   

            def column_generator(n, stop):
                while True:
                    if n > stop:
                        raise StopIteration
                    yield n
                    n += 1
            max_column = head_per_count_num

            ik = 4       # с 3  cтолбца       
            column_curr = raw_generator(ik, max_column+i)

    ##        print('context[get_chem_courier_objects_from_base].items() ====', context['get_chem_courier_objects_from_base'].items())

            the_val = 0
            excel_sheet['A1'] = tyresize_to_check_name
            num_counterrr = 0
            for key, value in context['get_chem_courier_objects_from_base'].items():
                num_counterrr += 1

                val = next(row_curr)
                excel_sheet['A2'] = '#'  
                excel_sheet.cell(row=val, column=1).value = num_counterrr         

                excel_sheet['B2'] = 'Бренд'
                excel_sheet.cell(row=val, column=2).value = key[0]
                excel_sheet.cell(row=val+1, column=2).value = 'ИТОГО'

                excel_sheet['C2'] = 'Получатель'
                excel_sheet.cell(row=val, column=3).value = key[1]

                the_val = val
    #        print('num_to_abc_convert_dict', num_to_abc_convert_dict)               #  !!!! DICT_T_T

            for column_n in columns_dict.values():

                col_num_list = []
                for nn in column_n:
                    val_column = next(column_curr)
                #    print('nn', nn)
                #    print('val_column', val_column)
                    col_num_list.append(val_column)
                excel_sheet[f'{num_to_abc_convert_dict.get(col_num_list[0])}2'] = 'Объем поставки, шт.'
                excel_sheet[f'{num_to_abc_convert_dict.get(col_num_list[1])}2'] = 'Объем поставки, USD'
                excel_sheet[f'{num_to_abc_convert_dict.get(col_num_list[2])}2'] = 'Средневзвешенная цена реализации единицы товара, USD'

            # заполнение данных в периодах
            list_val_in_per_per = []
    #        list_val_in_per = []
            for value in context['get_chem_courier_objects_from_base'].values():  
                for col in context['headers_len'].keys():
                    for tk, type_obj in value.items():
                        if tk == col:
                        #    temp_list = []
                            for tt in type_obj:
            #                    print('tt', tt)
                        #        temp_list.append(tt)
                                list_val_in_per_per.append(tt)
                        #    list_val_in_per.insert(len(list_val_in_per), temp_list)    
                        #    temp_list = []


            all_columns_names_period_reversed = list_val_in_per_per
            len_num = len(all_columns_names_period_reversed)
            some_dict = {}
            for column_num in range(1,len_num): 
                some_dict[column_num] = all_columns_names_period_reversed.pop(0) 

            # вставка значений в колонки по периодам (без итого)
            counter = 0     
            head_per_count_num = 4 + headers_len_periods_num 
            for row in range(3, the_val+1):                                                       
                for column in range(4,head_per_count_num):       
                    column_letter = get_column_letter(column)
                    counter = counter + 1
                    col_row_current_val = some_dict.get(counter)

                    excel_sheet[column_letter + str(row)] = col_row_current_val      # здесь вставляются данные 

            # END заполнение данных в периодах 
            # заполнение ИТОГО в периодах
            list_itogo_per = []
            for col in context['headers_len'].keys():
                for date, data_listt in context['itogo'].items(): 
                    if date == col:
                        for dt in data_listt:
                            #print('TTTTTcds', dt)
                            list_itogo_per.append(dt)
            itogo_sum_per_list = list(reversed(list_itogo_per))

            some_dict_sum_per = {}
            for column_num in range(4,head_per_count_num): 
                some_dict_sum_per[column_num] = itogo_sum_per_list.pop(-1) 

            #print('some_dict_sum_per', some_dict_sum_per)
            row_itogo_per = the_val+1
            for column_itogo_per in range(4,head_per_count_num):  
                excel_sheet.cell(row=row_itogo_per, column=column_itogo_per).value = some_dict_sum_per.get(column_itogo_per)   
            # END заполнение ИТОГО в периодах 

            lolumns_list_itogo = []
            count_to_three = 0
            counter = 0                                                     
            for column in range(head_per_count_num, head_per_count_num + 3):       
                column_letter = get_column_letter(column)
                count_to_three += 1
                lolumns_list_itogo.append(column_letter)
                if count_to_three == 3:
                    break
            all_columns_names_itogo_reversed = list(reversed(lolumns_list_itogo))
            num_to_abc_convert_itogo_dict = {}
            list_col_num = []
            for column_num in range(head_per_count_num, head_per_count_num + 3): 
                list_col_num.append(column_num)
                num_to_abc_convert_itogo_dict[column_num] = all_columns_names_itogo_reversed.pop() 


            itogo_colums_range_num = range(head_per_count_num, head_per_count_num + 3)
            excel_sheet[f'{num_to_abc_convert_itogo_dict.get(itogo_colums_range_num[0])}2'] = 'количество, шт.'   
            excel_sheet[f'{num_to_abc_convert_itogo_dict.get(itogo_colums_range_num[1])}2'] = 'сумма импорта долл.США без НДС'  
            excel_sheet[f'{num_to_abc_convert_itogo_dict.get(itogo_colums_range_num[2])}2'] = 'средневзвешенная цена, долл. США без НДС'  


            def raw_generator_sum(n, stop):
                while True:
                    if n > stop:
                        raise StopIteration
                    yield n
                    n += 1
            max_raw = row_value + 2  # (+1 строка для ИТОГО и записи что согласно хим курьеру)
            i = 3       # с 3  строки        
            row_curr_sum = raw_generator_sum(i, max_raw+i)

            list_itogo_sum_sum_val = context['num_summ_right_columns_result_itog_sum']

            last_row_in_table_num = 0
            for key, vall in context['get_chem_courier_objects_from_base'].items():
                val = next(row_curr_sum)
                for res_key, res_val in context['num_summ_right_columns_result'].items(): 
                    if res_key == key:
                        excel_sheet.cell(row=val, column=list_col_num[0]).value = res_val[0]
                        excel_sheet.cell(row=val, column=list_col_num[1]).value = res_val[1]
                        excel_sheet.cell(row=val, column=list_col_num[2]).value = res_val[2]

                        excel_sheet.cell(row=val+1 , column=list_col_num[0]).value = list_itogo_sum_sum_val[0] # ДЛЯ ИТОГО СУММА ВСЕХ СУММ
                        excel_sheet.cell(row=val+1 , column=list_col_num[1]).value = list_itogo_sum_sum_val[1]
                        excel_sheet.cell(row=val+1 , column=list_col_num[2]).value = list_itogo_sum_sum_val[2]
                last_row_in_table_num = val +1

            excel_sheet.cell(row=last_row_in_table_num+2, column=1).value = 'Согласно данным информационно-аналитическго агентства «Хим-Курьер»'   #
            #  Для отрисовки лийний у таблицы:
            from openpyxl.styles.borders import Border, Side
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
    #           excel_sheet.cell(row=3, column=2).border = thin_border
            #for i in excel_sheet['A1':'D8']:
            cell_range = excel_sheet['A2':f'{num_to_abc_convert_itogo_dict.get(itogo_colums_range_num[2])}{last_row_in_table_num}']
            for row in cell_range:
                for cell in row:
                    cell.border = thin_border

            #wb.save("ChemcourierProgr.xlsx")
                #wb.save('/home/user/Desktop/FileName.xlsx')

            # сохранение в папку DOWNLOADS пользователя:  
            user_path_is = os.path.expanduser("~")
            #print('path_is', user_path_is)
            download_folder_exist = user_path_is + "/Downloads/"                            # создать файл в папке Downloads
            if os.path.exists(download_folder_exist):
                users_download_path = user_path_is + "/Downloads/" + 'ChemcourierProgr.xlsx'
                dowload_to = os.path.normpath(users_download_path)     
                print('AAAA')
            else:                                                                           # если папки Downloads нет - создать файл в корневой папке пользователя
                users_download_path = user_path_is  + 'ChemcourierProgr.xlsx'
                dowload_to = os.path.normpath(users_download_path)
            
            
            def not_in_use(filename):                                               # проверка - используется ли в данный момент файл (если он откурыт -то будет попытка его переписыть с ошибкой - надо избеать)
                    try:
                        os.rename(filename,filename)
                        return True
                    except:  
                        return False     
            if os.path.exists(dowload_to):          
                if not_in_use(dowload_to):
                    wb.save(dowload_to)
            else:
                wb.save(dowload_to)
            #wb.save(dowload_to)

            ####### END СКАЧИВАНИЕ ФАЙЛА - EXCEL ТАБЛИЦА с данными из таблицы
            wb.close()

        return context  
    
class ChemcourierTableProgressiveModelUpdateView(View):

    def post(self, request):
        models.CHEM_PNJ_IN_TABLE_LIST = []
        models.CHEM_TABLE_FINAL_DATA = {}
        models.CHEM_TABLE_FINAL_DATA_FINAL = {}

        #print(request.POST, 'TTTH')
        # 1. получаем период от пользователя для поиска и установки initial значения в дате формы выбора
        get_period_start = request.POST.get('periods_start') 
        if get_period_start:
            forms.INITIAL_PERIOD_START = get_period_start
        get_period_end = request.POST.get('periods_end') 
        if get_period_end:
            forms.INITIAL_PERIOD_END = get_period_end
        # 2. получаем типоразмер от пользователя для поиска и установки initial значения в типоразмере формы выбора    
        get_tyresize = request.POST.get('tyresizes') 
        if get_tyresize:
            forms.INITIAL_TYREISIZE = get_tyresize

        # 3. получаем бренд от пользователя для поиска и установки initial значения в бренде формы выбора    
        get_tyrebrand = request.POST.get('tyrebrands') 
        if get_tyrebrand and get_tyrebrand != '-':
            forms.INITIAL_BRANDS = get_tyrebrand
        else:
            forms.INITIAL_BRANDS = None

        # 4. получаем получателя от пользователя для поиска и установки initial значения в получателях формы выбора    
        get_reciever = request.POST.get('recievers') 
        if get_reciever and get_reciever != '-':
            forms.INITIAL_RECIEVER = get_reciever
        else:
            forms.INITIAL_RECIEVER = None

        # 5. получаем страну производства от пользователя для поиска и установки initial значения в странах производства формы выбора    
        get_prod_countrys = request.POST.get('prod_countrys') 
        if get_prod_countrys and get_prod_countrys != '-':
            forms.INITIAL_PRODCOUTRYS = get_prod_countrys
        else:
            forms.INITIAL_PRODCOUTRYS = None

        # 6. получаем группу производства от пользователя для поиска и установки initial значения в группах формы выбора    
        get_prod_groups = request.POST.get('prod_groups') 
        if get_prod_groups and get_prod_groups != '-':
            forms.INITIAL_GROUPS = get_prod_groups
        else:
            forms.INITIAL_GROUPS = None

        # 7. курс валют на дату
        chosen_date_for_currency_year = request.POST.getlist('chosen_date_for_currency_year') 
        chosen_date_for_currency_month = request.POST.getlist('chosen_date_for_currency_month') 
        chosen_date_for_currency_day = request.POST.getlist('chosen_date_for_currency_day') 
        chosen_date_for_currency = chosen_date_for_currency_year + chosen_date_for_currency_month + chosen_date_for_currency_day
        if chosen_date_for_currency:
            chosen_date_for_currency = '-'.join(str(x) for x in chosen_date_for_currency)
            check_date = datetime.datetime.strptime(chosen_date_for_currency, "%Y-%m-%d").date()        #  если пользователем введена дана превышающая текущую для получения курса валют то нао скинуть на сегодня:
            if check_date > datetime.datetime.now().date():
                prices_models.CURRENCY_ON_DATE is False
            else:
                prices_models.CURRENCY_DATE_GOT_FROM_USER = chosen_date_for_currency
                prices_models.CURRENCY_ON_DATE is True

        # 8. создать EXCEL CHEMCOURIER
        chemcuorier_download = request.POST.getlist('chemcuorier_progr_download') 
        if chemcuorier_download:
            forms.CHEMCOURIER_PROGRESSIVE_EXCEL_CREATE = True

        return HttpResponseRedirect(reverse_lazy('chemcurier:chemcurier_table_progressive'))