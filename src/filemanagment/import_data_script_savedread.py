import datetime
from enum import unique
from operator import index
#from turtle import st
from unicodedata import decimal, name
from webbrowser import get
from django.shortcuts import render
from django.views.generic import FormView, TemplateView
from . import forms
from django.http import HttpResponseRedirect
from django.urls import reverse_lazy
import openpyxl 
import openpyxl.utils.cell
import re
from dictionaries import models as dictionaries_models
from tyres import models as tyres_models
from sales import models as sales_models
from itertools import count, islice
from abc_table_xyz import models as abc_table_xyz_models
from datetime import datetime
from prices import models as prices_models
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import time
import asyncio 
import threading
import os
from openpyxl import load_workbook
from django.core.files.storage import default_storage


# создать временный файл
def make_a_copy_of_users_chemc_file(file_to_read, sheet):

    wb1 = file_to_read
    wb1.save("PEMANENT_FILE.xlsx")
    #print('создана копия файла - FILE CREATED')

    return wb1

# запись в базу данных
def chem_courier_bulk_write_ib_bd(dict_of_data):
    
    #    print('ЗАПИСЬ В БАЗУ ХИМ ПАЧЫНАЕЦЦА', datetime.now())
        MAIN_chemcirier_import_dict = dict_of_data
    #    print('LEN DICT', len(MAIN_chemcirier_import_dict))
        # 1.1 создать объекты продажа на дату:
        che_curier_obj_tyre_bulk_list = []
        currency_chem = dictionaries_models.Currency.objects.get_or_create(currency = 'USD')[0]
        for key, val in MAIN_chemcirier_import_dict.items():
        #    print(key[0], key, 'CHEM',  val[0][0])
    #    # 1.2 создать объекты типоразмер производитель группа:  
        #    print('tyre_size_chem', key[0])
        #    print('producer_chem', key[2])
        #    print('group_chem', key[3])
        #    print('recipipient_namee', key[4])
        #    print('recipipient_namee', key[5])
            list_val_price_all_periods_for_row = val[0]
            for data_month_chem_val in list_val_price_all_periods_for_row:
                if data_month_chem_val[1] and data_month_chem_val[2]:           # если если есть шт. и сумма на дату - создать объект:
        #            print('data_month_chem', data_month_chem_val[0])
        #            print('val_on_moth_che', data_month_chem_val[1])
        #            print('money_on_moth_chem ', data_month_chem_val[2])
                    group_chem_obect = None
                    if key[3] == 'Шиныдлястроительнойипромышленнойтехники' or key[3] == 'ШиныдлягрузовыхавтоЦМК':
                        group_chem_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='грузовые')
                    elif key[3] == 'Шиныдлялегковыхавто':
                        group_chem_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='легковые')
                    elif key[3] == 'Шиныдляс/хтехники':
                        group_chem_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='с/х')
                    elif key[3] == 'Шиныдлялегкогрузовыхавто':
                        group_chem_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='легкогруз')      
                # РАБОЧИЙ ВАРИАНТ _ НО МЕДЛЕННЫЙ _ ЗАПИСЬ В БАЗУ ДАННЫХ    
                ######    che_curier_obj_tyre = prices_models.ChemCurierTyresModel.objects.update_or_create(
                ######        tyre_size_chem = key[0],
                ######        producer_chem = key[2],
                ######        # #"'ШиныдлягрузовыхавтоЦМК   'Шиныдлялегковыхавто' 'Шиныдлялегкогрузовыхавто' 'Шиныдляс/хтехники' 'Шиныдлястроительнойипромышленнойтехники'
                ######        group_chem = group_chem_obect,
                ######        reciever_chem = key[4],
                ######        prod_country = key[5],
                ######        currency_chem = currency_chem,
                ######        data_month_chem = data_month_chem_val[0],
                ######        val_on_moth_chem = data_month_chem_val[1],
                ######        money_on_moth_chem = data_month_chem_val[2], 
                ######        average_price_in_usd = data_month_chem_val[3],
                ######    #    month_counter = val[1] 
                ######        #price_val_money_data = prices_models.DataPriceValMoneyChemCurierModel.objects.filter(data_month_chem__date=date_month)          #abc_obj_set[0].sales.add(sales_obj)
                ######    )[0]
                # END РАБОЧИЙ ВАРИАНТ _ НО МЕДЛЕННЫЙ _ ЗАПИСЬ В БАЗУ ДАННЫХ    
        
        # BULK_CREATE
                    # ПРЕдварительно проверка - есть ли уже в базе данный объекту
                    chem_obj_not_exist_alredy = True
                    try:
                        prices_models.ChemCurierTyresModel.objects.get(
                        tyre_size_chem = key[0],
                        producer_chem = key[2],
                        # #"'ШиныдлягрузовыхавтоЦМК   'Шиныдлялегковыхавто' 'Шиныдлялегкогрузовыхавто' 'Шиныдляс/хтехники' 'Шиныдлястроительнойипромышленнойтехники'
                        group_chem = group_chem_obect,
                        reciever_chem = key[4],
                        prod_country = key[5],
                        currency_chem = currency_chem,
                        data_month_chem = data_month_chem_val[0],
                        val_on_moth_chem = data_month_chem_val[1],
                        money_on_moth_chem = data_month_chem_val[2], 
                        average_price_in_usd = data_month_chem_val[3],                                
                        )
                    except:
                        chem_obj_not_exist_alredy = False
                    if chem_obj_not_exist_alredy is False:      # если объект в базе отстутствует - добавляем в список для bulk_create:
                    # END ПРЕдварительно проверка - есть ли уже в базе данный объекту
                        che_curier_obj_tyre_bulk_list.append(prices_models.ChemCurierTyresModel(
                            tyre_size_chem = key[0],
                            producer_chem = key[2],
                            # #"'ШиныдлягрузовыхавтоЦМК   'Шиныдлялегковыхавто' 'Шиныдлялегкогрузовыхавто' 'Шиныдляс/хтехники' 'Шиныдлястроительнойипромышленнойтехники'
                            group_chem = group_chem_obect,
                            reciever_chem = key[4],
                            prod_country = key[5],
                            currency_chem = currency_chem,
                            data_month_chem = data_month_chem_val[0],
                            val_on_moth_chem = data_month_chem_val[1],
                            money_on_moth_chem = data_month_chem_val[2], 
                            average_price_in_usd = data_month_chem_val[3],))
                    
        bulk_msj = prices_models.ChemCurierTyresModel.objects.bulk_create(che_curier_obj_tyre_bulk_list)

        #temporary_created_file = os.path.abspath("PEMANENT_FILE.xlsx")
        #print('PATH', temporary_created_file)
        #os.remove("PEMANENT_FILE.xlsx")

            
        return print('ЗАПИСЬ В БАЗУ ХИМ ЗАВЕРШЕНА', datetime.now())
      

# удалить временный файл        
def delete_temp_file():
    try:
        temporary_created_file = os.path.abspath("PEMANENT_FILE.xlsx")
        #print('PATH', temporary_created_file)
        os.remove("PEMANENT_FILE.xlsx")
        print('временный файл удален')
    except:
        pass


# расчет объема для считывания из файла по N строк
def rows_in_file_limiter(copy_file, list_of_sheet_potential_names_var_list):    
    file_to_read = copy_file
    sheet = None
    list_chemcurier_years = range(11, 50)
    for year_d in list_chemcurier_years:
        for sheet_name in list_of_sheet_potential_names_var_list:
            try:
                sheet = file_to_read[f'{sheet_name} 20{year_d}']
                #sheet = file_to_read['ИМПОРТ 2022']    # Читаем файл и лист1 книги excel 
            except:
                pass
    if sheet:       # РАЗБИТИЕ EXCEL НА ЧАСТИ ДЛЯ СЧИТЫВАНИЯ:
        max_row = sheet.max_row
        #print('sheet.max_row', sheet.max_row) 
        SSTEP = 3000      # шаг cсчитывания с excel
        big_steps_num = int(max_row / SSTEP)
        small_step = max_row % SSTEP
        #print('list_of_cycles', big_steps_num, 'ttttt', small_step, '@@@')
        excel_red_cycles_list = []
        st_row = 1
        end_row = 0
        for n in range(0, big_steps_num):  #1, 2, 
            end_row += SSTEP
            cyycle = st_row, end_row
            excel_red_cycles_list.append(cyycle)
            st_row = end_row + 1
            if n == big_steps_num-1 and end_row + 1 < end_row + small_step:
                cyycle = end_row + 1, end_row + small_step
                excel_red_cycles_list.append(cyycle)
        if not excel_red_cycles_list:
            cyycle = 1, small_step
            excel_red_cycles_list.append(cyycle)


        ####    e = create_engine('sqlite:///sqlite3.db', pool_recycle=39600) 
        ####    c = e.connect()
        ###    from sqlalchemy import create_engine
        ####    from django.db import connection

        # проход по документу:
        for cccycle_ten_thosand in excel_red_cycles_list: 
            #try:
                #read_from_chem_courier_copy_file(copy_file, list_of_sheet_potential_names_var_list, some_func, cccycle_ten_thosand)
            read_result = read_from_chem_courier_copy_file(copy_file, list_of_sheet_potential_names_var_list, cccycle_ten_thosand)
            chem_courier_bulk_write_ib_bd(read_result)
            #except:
            cccycle_ten_thosand = cccycle_ten_thosand
        ####    connection.close()

        # удалить временный файл:
        delete_temp_file()

    #return print('###', excel_red_cycles_list)
    return print('')
          

# считывание  с временного файла
#def read_from_chem_courier_copy_file(copy_file, list_of_sheet_potential_names_var_list, some_func, step_to_read):
def read_from_chem_courier_copy_file(copy_file, list_of_sheet_potential_names_var_list, step_to_read):
    
    brend_chemcurier_dict = {} # словарь, в который закидываются данные о действующих бренде ХИМКУРЬЕР
    recipipient_chemcurier_dict = {} # словарь, в который закидываются данные о получателе. на дату ХИМКУРЬЕР
    prod_country_column_dict = {} # словарь, в который закидываются данные о стране-производителе. на дату ХИМКУРЬЕР
    tiresize_chemcurier_dict = {} # словарь, в который закидываются данные о действующих типоразмере ХИМКУРЬЕР
    tire_group_chemcurier_dict = {} # словарь, в который закидываются данные о группе шин ХИМКУРЬЕР
    pieces_month_chemcurier_dict = {} # словарь, в который закидываются данные о шт. на дату ХИМКУРЬЕР
    money_month_chemcurier_dict = {} # словарь, в который закидываются данные о шт. на дату ХИМКУРЬЕР
    MAIN_chemcirier_import_dict = {}    # главгый словарь вкуладки импорт ХИМКУРЬЕР

#    file_to_read = copy_file

    ### корявая вставка - но на отткытие -закрытие файла
    #temporary_created_file = os.path.abspath("PEMANENT_FILE.xlsx")
    created2 = load_workbook(filename = 'PEMANENT_FILE.xlsx')
    file_to_read = created2
    ######


    list_of_sheets = list_of_sheet_potential_names_var_list
    sheet = None
    list_chemcurier_years = range(11, 50)
    for year_d in list_chemcurier_years:
        for sheet_name in list_of_sheet_potential_names_var_list:
            try:
            #    sheet = file_to_read[f'{sheet_name} 20{year_d}']
                sheet =created2[f'{sheet_name} 20{year_d}']

                #
            except:
                pass
    if sheet:   

    ############ END создать физическую копию файла юзера для работы с ней
    #    print(';******8888_____+')
        

        counter_limiter_counter = 0       # счетчик считывания для формирования словаря для записи до 1000 строк
        #print('!pp!', cell.row)

        chemcirier_rows_counter = []
        brand_names_list = ['бренд', 'Бренд'] 
        reciever_names_list = ['получатель', 'Получатель'] 
        countr_names_list = ['страна производства', 'Страна производства']
        trsz_names_list = ['типоразмер', 'Типоразмер']
        group_names_list = ['подгруппа', 'Подгруппа']   
        



        # модификация
        step_to_read = step_to_read
    #    print('KKKYY', step_to_read[0], step_to_read[1])

        #mmmax_row = sheet.max_row
        
        mmmin_row = step_to_read[0]
        mmmax_row = step_to_read[1]
        
        for row in sheet.rows:  # БЫСТРЫЙ СПОСОБ
        #for row in sheet.iter_rows(step_to_read[0], step_to_read[1]):
        # END модификация    


        #    print('ROW iss', row )        
            for cell in row: 
                # 1 Парсинг 
                if cell.value in brand_names_list:          # получаем колонку 'бренд'  ХИМКУРЬЕР
                    brand_column = cell.column
                    brand_row = cell.row

                    if step_to_read[0] == 1:
                        mmmin_row = step_to_read[0]
                    else:
                        mmmin_row = step_to_read[0] - brand_row


                    #for col in sheet.iter_cols(min_row=brand_row+1, min_col=brand_column, max_col=brand_column, max_row=mmmax_row):
                    for col in sheet.iter_cols(min_row=brand_row+mmmin_row, min_col=brand_column, max_col=brand_column, max_row=mmmax_row):                    
                        for cell in col:                            
                            brend_chemcurier_dict[cell.row] = cell.value

                if cell.value in reciever_names_list:          # получаем колонку 'получатель'  ХИМКУРЬЕР
                    recipient_column = cell.column
                    brand_row = cell.row

                    if step_to_read[0] == 1:
                        mmmin_row = step_to_read[0]
                    else:
                        mmmin_row = step_to_read[0] - brand_row

                    #for col in sheet.iter_cols(min_row=brand_row+1, min_col=recipient_column, max_col=recipient_column, max_row=sheet.max_row):
                    for col in sheet.iter_cols(min_row=brand_row+mmmin_row, min_col=recipient_column, max_col=recipient_column, max_row=mmmax_row):    
                        for cell in col:                            
                            recipipient_chemcurier_dict[cell.row] = cell.value        

                if cell.value in countr_names_list:          # получаем колонку 'страна производства'  ХИМКУРЬЕР
                    prod_country_column = cell.column
                    brand_row = cell.row

                    if step_to_read[0] == 1:
                        mmmin_row = step_to_read[0]
                    else:
                        mmmin_row = step_to_read[0] - brand_row

                    #for col in sheet.iter_cols(min_row=brand_row+1, min_col=prod_country_column, max_col=prod_country_column, max_row=sheet.max_row):
                    for col in sheet.iter_cols(min_row=brand_row+mmmin_row, min_col=prod_country_column, max_col=prod_country_column, max_row=mmmax_row):    
                        for cell in col:                            
                            prod_country_column_dict[cell.row] = cell.value   

                if cell.value in trsz_names_list:          # получаем колонку 'типоразмер' ХИМКУРЬЕР
                    tyr_group_column = cell.column
                    tyr_group_row = cell.row


                    if step_to_read[0] == 1:
                        mmmin_row = step_to_read[0]
                    else:
                        mmmin_row = step_to_read[0] - tyr_group_row

                    #for col in sheet.iter_cols(min_row=tyr_group_row+1, min_col=tyr_group_column, max_col=tyr_group_column, max_row=sheet.max_row):
                    for col in sheet.iter_cols(min_row=tyr_group_row+mmmin_row, min_col=tyr_group_column, max_col=tyr_group_column, max_row=mmmax_row):    
                        for cell in col: 
                            if cell.value:                          
                                tiresize_chemcurier_dict[cell.row] = cell.value.replace(' ', '')
                                
                if cell.value in group_names_list:          # получаем колонку 'подгруппа' ХИМКУРЬЕР
                    season_column = cell.column
                    season_row = cell.row

                    if step_to_read[0] == 1:
                        mmmin_row = step_to_read[0]
                    else:
                        mmmin_row = step_to_read[0] - season_row

                    #for col in sheet.iter_cols(min_row=season_row+1, min_col=season_column, max_col=season_column, max_row=sheet.max_row):
                    for col in sheet.iter_cols(min_row=season_row+mmmin_row, min_col=season_column, max_col=season_column, max_row=mmmax_row):
                        for cell in col:
                            if cell.value:                                          
                                tire_group_chemcurier_dict[cell.row] = cell.value.replace(' ', '')
                #### парсим штуки деньги из объединенных ячеек:
                stuck_names_list = ['штук', 'тыс. штук', 'Штук', 'Тыс. Штук', 'тысяч штук', 'шт.', 'шт', 'тыс. шт.' ]
                doll_names_list_list = ['долларов*', 'долларов', 'тыс. долларов*', 'тыс. долларов', 'Долларов*', 'Долларов', 'Тыс. долларов', 'Тыс. Долларов', 'тыс. долл.', 'Тыс. долл', 'тысяч долларов', 'тыс. долл',]
                for merged_cell in sheet.merged_cells.ranges:           # для объединенных ячеек их поиск
                    if cell.coordinate in merged_cell:
                        #print('cell.coordinate', cell.coordinate)  
                        if cell.value in stuck_names_list:
                            stucki_coordiantes_merged = str(merged_cell).split(':')
                            if stucki_coordiantes_merged:
                                # получим начальные и конечные координаты ячеек - дат под заголовком штуки:
                                start_coord = stucki_coordiantes_merged[0]
                                end_coord = stucki_coordiantes_merged[1]
                                # переводим координаты столбцов из буквенных в циферные значения:
                                numbered_start_coord = coordinate_from_string(start_coord)
                                numbered_end_coord = coordinate_from_string(end_coord)
                                rowwww = numbered_start_coord[1] # rowwww  == rowwww1 = numbered_end_coord[1]   строка то ведь одна и та же
                                colll_start_num = column_index_from_string(numbered_start_coord[0])
                                colll_end_num = column_index_from_string(numbered_end_coord[0])
                                #print('rowwww', rowwww, 'colll_start_num = ', colll_start_num, 'colll_end_num =', colll_end_num)
                                #for col in sheet.iter_cols(min_row=rowwww+1, max_col=colll_end_num, max_row=rowwww+1, min_col=colll_start_num):
                                for col in sheet.iter_cols(min_row=rowwww+1, max_col=colll_end_num, max_row=mmmax_row, min_col=colll_start_num):
                                    for cell in col:
                                        #print(cell, '5656767756')
                                        if cell.value:                                                                              # если ячейка с датой: проверка и получение даты
                                            is_date = (isinstance(cell.value, datetime))
                                            if is_date == True:
                                                month_chemcurier = cell.value.date()
                                                #print(month_chemcurier)
                                                month_chemcurier_column = cell.column
                                                month_chemcurier_row = cell.row

                                                if step_to_read[0] == 1:
                                                    mmmin_row = step_to_read[0]
                                                else:
                                                    mmmin_row = step_to_read[0] - month_chemcurier_row

                                                #for col in sheet.iter_cols(min_row=month_chemcurier_row+1, min_col=month_chemcurier_column, max_col=month_chemcurier_column, max_row=sheet.max_row):
                                                for col in sheet.iter_cols(min_row=month_chemcurier_row+mmmin_row, min_col=month_chemcurier_column, max_col=month_chemcurier_column, max_row=mmmax_row):
                                                    list_col_values = []
                                                    for cell in col:   
                                                        #print(cell.value, 'VV')   
                                                        column_value_row = cell.value, cell.row 
                                                        list_col_values.append(column_value_row)
                                                    pieces_month_chemcurier_dict[month_chemcurier] = list_col_values
                                                #for roww in sheet.iter_cols(min_row=month_chemcurier_row+1, min_col=month_chemcurier_column, max_col=month_chemcurier_column, max_row=sheet.max_row):
                                                #    list_col_values = []
                                                #    for cell in roww:   
                                                #        #print(cell.value, 'VV')   
                                                #        column_value_row = cell.value, month_chemcurier
                                                #        list_col_values.append(column_value_row)
                                                #        #print(list_col_values)
                                                #    pieces_month_chemcurier_dict[cell.row] = list_col_values 

                        if cell.value in doll_names_list_list:
                            stucki_coordiantes_merged = str(merged_cell).split(':')
                            if stucki_coordiantes_merged:
                                # получим начальные и конечные координаты ячеек - дат под заголовком штуки:
                                start_coord = stucki_coordiantes_merged[0]
                                end_coord = stucki_coordiantes_merged[1]
                                # переводим координаты столбцов из буквенных в циферные значения:
                                numbered_start_coord = coordinate_from_string(start_coord)
                                numbered_end_coord = coordinate_from_string(end_coord)
                                rowwww = numbered_start_coord[1] # rowwww  == rowwww1 = numbered_end_coord[1]   строка то ведь одна и та же
                                colll_start_num = column_index_from_string(numbered_start_coord[0])
                                colll_end_num = column_index_from_string(numbered_end_coord[0])
                                #print('rowwww', rowwww, 'colll_start_num = ', colll_start_num, 'colll_end_num =', colll_end_num)

                                #for col in sheet.iter_cols(min_row=rowwww+1, max_col=colll_end_num, max_row=rowwww+1, min_col=colll_start_num):
                                for col in sheet.iter_cols(min_row=rowwww+1, max_col=colll_end_num, max_row=step_to_read[1], min_col=colll_start_num):
                                    for cell in col:
                                        #print(cell, '5656767756')
                                        if cell.value:                                                                              # если ячейка с датой: проверка и получение даты
                                            is_date = (isinstance(cell.value, datetime))
                                            if is_date == True:
                                                month_chemcurier = cell.value.date()
                                                #print(month_chemcurier)
                                                month_chemcurier_column = cell.column
                                                month_chemcurier_row = cell.row

                                                if step_to_read[0] == 1:
                                                    mmmin_row = step_to_read[0]
                                                else:
                                                    mmmin_row = step_to_read[0] - month_chemcurier_row

                                                #for col in sheet.iter_cols(min_row=month_chemcurier_row+1, min_col=month_chemcurier_column, max_col=month_chemcurier_column, max_row=sheet.max_row):
                                                for col in sheet.iter_cols(min_row=month_chemcurier_row+mmmin_row, min_col=month_chemcurier_column, max_col=month_chemcurier_column, max_row=step_to_read[1]):                           
                                                    chemcirier_rows_counter = []
                                                    list_col_values = []
                                                    for cell in col:   
                                                        #print(cell.value, 'VV') 
                                                        chemcirier_rows_counter.append(cell.row) 
                                                        column_value_row = cell.value, cell.row 
                                                        list_col_values.append(column_value_row)
                                                    money_month_chemcurier_dict[month_chemcurier] = list_col_values

        #### ХИМКУРЬЕР Ч 2
                       #### пересборка словаря pieces_month_chemcurier_dict ключ - номер строки, значения - дата и значение шт.     
        #1.1 пересборка словарей pieces_month_chemcurier_dict и  money_month_chemcurier_dict:               # 2022-01-01 [(28, 8), (None, 9), (180, 10), (2426, 11), (24, 12), (17, 13), (6, 14), (89, 15)] PPP
        if chemcirier_rows_counter:
        #    print('+++', chemcirier_rows_counter)
            min_row_value = min(chemcirier_rows_counter)                                                        # 2022-02-01 [(36, 8), (960, 9), (None, 10), (None, 11), (179, 12), (None, 13), (4, 14), (156, 15)] PPP
            max_row_value = max(chemcirier_rows_counter)
            #print('min_row_value', min_row_value)
            #print('max_row_value', max_row_value)
            chemcirier_rows_counter = [min_row_value, max_row_value]

            new_pieces_month_chemcurier_dict ={}
            new_money_month_chemcurier_dict ={}
            list_col_values_len = max_row_value - min_row_value
            for str_n in range(chemcirier_rows_counter[0], chemcirier_rows_counter[1]+1):
                list_sates_values_on_row = []
                for k, v in pieces_month_chemcurier_dict.items():
            #        print('!!!', k, v)
                    for list_iter in range(0, list_col_values_len+1):
                        #print(k, v[list_iter][1])
                        if v[list_iter][1] == str_n:
                            cell_val = k, v[list_iter][0]
                            #print(cell_val, 'OOOO')
                            list_sates_values_on_row.append(cell_val)
            #            print('===',list_sates_values_on_row)
                new_pieces_month_chemcurier_dict[str_n] = list_sates_values_on_row
                list_sates_values_on_row1 = []
                for k, v in money_month_chemcurier_dict.items():
                    for list_iter in range(0, list_col_values_len+1):
                        #print(k, v[list_iter][1])
                        if v[list_iter][1] == str_n:
                            cell_val = k, v[list_iter][0]
                            #print(cell_val, 'OOOO')
                            list_sates_values_on_row1.append(cell_val)
                        #print(list_sates_values_on_row1)
                new_money_month_chemcurier_dict[str_n] = list_sates_values_on_row1

            #print(new_money_month_chemcurier_dict)
        #    for k, v in new_money_month_chemcurier_dict.items():
        #        print('-----', k, v)
            ################ ФОРМИРОВАНИЕ ЕДИНОГО СЛОВАРЯ ПО СПАРСЕННЫМ ДАННЫМ ХИМКУРЬЕР ВКЛАДКИ ИМПОРТ:
            
            for row_num, tyr_size in tiresize_chemcurier_dict.items():
                brand_namee = brend_chemcurier_dict.get(row_num, 0)
                group_namee = tire_group_chemcurier_dict.get(row_num, 0)
                recipipient_namee = recipipient_chemcurier_dict.get(row_num, 0)
                prod_country_namee = prod_country_column_dict.get(row_num, 0)
                #print('LEN', len(new_money_month_chemcurier_dict.get(row_num, 1)),' ===== ')
                periods_counter_max_val = len(new_money_month_chemcurier_dict.get(row_num, 1)) 
                periods_counter = 0
                values_prices_per_month_list = []
            #    print('periods_counter]!!!!!', periods_counter)
                if periods_counter_max_val: # если есть периоды
                    while periods_counter < periods_counter_max_val:   
            #            print('t7t7t7', new_pieces_month_chemcurier_dict.get(row_num, 0)[periods_counter])                                        
                        pices_per_month = new_pieces_month_chemcurier_dict.get(row_num, 0)[periods_counter][1]
            #            print('===== 1', pices_per_month)
                        money_per_month = new_money_month_chemcurier_dict.get(row_num, 0)[periods_counter][1]
            #            print('===== 2',money_per_month)
                        period_month = new_money_month_chemcurier_dict.get(row_num, 0)[periods_counter][0]                            
                        if money_per_month is not None and pices_per_month is not None:
                            average_price_in_usd_calc = money_per_month / pices_per_month 
                        else:
                            average_price_in_usd_calc = None
                        pices_per_month_money_per_month = period_month, pices_per_month, money_per_month, average_price_in_usd_calc
                        periods_counter += 1 
                        values_prices_per_month_list.append(pices_per_month_money_per_month)
                
                #MAIN_chemcirier_import_dict[tyr_size, row_num] = brand_namee, group_namee, pices_per_month, money_per_month 
                MAIN_chemcirier_import_dict[tyr_size, row_num, brand_namee, group_namee, recipipient_namee, prod_country_namee] = list(values_prices_per_month_list), periods_counter 
        
        # перечень всех доступных дат (месяцев):
        month_in_chemcurier_table = list(money_month_chemcurier_dict.keys())
        #print('month_in_chemcurier_table', month_in_chemcurier_table)

    #print('MAIN_chemcirier_import_dict', MAIN_chemcirier_import_dict)
        
    file_to_read.close()  

    #MAIN_chemcirier_import_dict_is = MAIN_chemcirier_import_dict
    #MAIN_chemcirier_import_dict = {} # обнулить

    #some_func(MAIN_chemcirier_import_dict_is)
    #print('LLEN', len(MAIN_chemcirier_import_dict_is))
    #chem_courier_bulk_write_ib_bd(MAIN_chemcirier_import_dict_is)


    return MAIN_chemcirier_import_dict


#async def read_from_file(self):
def read_from_file():
  
    #print('ПРОВЕРКА ФАЙЛА - выбор форма а либо b/ b химкурьер ',  datetime.now())


    row_value = int
    tyresize_list = []
    tyremodel_list = []
    tyreparametrs_list = [] 
    sales_list = []                 #объем продаж на дату
    planned_costs = []              #плановая себестоимость/плановые затраты planned_costs 
    semi_variable_costs = []        #плановая себестоимость
    belarus902price_costs = []      #прейскуранты №№9, 902
    tpsrussiafcaprice_costs = []    #ТПС РФ FCA
    tpskazfcaprice_costs = []       #ТПС Казахстан FCA
    tpsmiddleasiafcaprice_costs = [] #ТПС Азия
    current_prices = []   # Действующие цены
    contragent_list = []            #контрагент
    column_sell_date = ''           #строка с датой продажи
    tyretype_row_dict = {}          # словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры = типоразмер
    model_row_dict = {}             # словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры = модель
    params_row_dict = {}            # словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры = параметры
    saless_row_dict = {}            # словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры = объем продаж
    planned_costs_row_dict = {}     # словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры = плановая себестоимость
    semi_variable_costs_row_dict = {} # словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры  = прямые затраты
    belarus902price_costs_row_dict = {}# словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры  = прейскуранты №№9, 902
    tpsrussiafcaprice_costs_row = {}# словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры  = ТПС РФ FCA
    tpskazfcaprice_cost_row = {}# словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры  = ТПС Казахстан FCA
    tpsmiddleasiafcaprice_costs_row = {}# словарь, в который закидываются данные из строк построчно. ключ - номер строки, параметры  = ТПС Средняя Азия, Закавказье, Молдова FCA
    current_prices_row = {}                 # словарь, в который закидываются данные о действующих ценах
    indexes_row_dict = {}                # словарь индексы ДОПОЛНИТЕЛЬНЫЕ, в который закидываются данные об индексах. ключ - номер строки, параметры = индексы
    season_row_dict = {}            # словарь сезонность ДОПОЛНИТЕЛЬНЫЕ, в который закидываются данные об сезонность. ключ - номер строки, параметры = сезонность
    thread_row_dict = {}            # словарь рисунок протектора ДОПОЛНИТЕЛЬНЫЕ, в который закидываются данные об рисунок протектора. ключ - номер строки, параметры = рисунок протектора
    ax_row_dict = {}            # словарь ось ДОПОЛНИТЕЛЬНЫЕ, в который закидываются данные об ось. ключ - номер строки, параметры = ось
    usability_row_dict = {}            # словарь применяемость ДОПОЛНИТЕЛЬНЫЕ, в который закидываются данные об применяемость. ключ - номер строки, параметры = применяемость
    row_parsing_sales_costs_prices_dict = {}     # словарь, в который закидываются данные из строк построчно. ключ - шина, параметры = данные о продажах, минималках и прайсах
    brend_chemcurier_dict = {} # словарь, в который закидываются данные о действующих бренде ХИМКУРЬЕР
    tiresize_chemcurier_dict = {} # словарь, в который закидываются данные о действующих типоразмере ХИМКУРЬЕР
    tire_group_chemcurier_dict = {} # словарь, в который закидываются данные о группе шин ХИМКУРЬЕР
    pieces_month_chemcurier_dict = {} # словарь, в который закидываются данные о шт. на дату ХИМКУРЬЕР
    money_month_chemcurier_dict = {} # словарь, в который закидываются данные о шт. на дату ХИМКУРЬЕР
    recipipient_chemcurier_dict = {} # словарь, в который закидываются данные о получателе. на дату ХИМКУРЬЕР
    prod_country_column_dict = {} # словарь, в который закидываются данные о стране-производителе. на дату ХИМКУРЬЕР
    MAIN_chemcirier_import_dict = {}    # главгый словарь вкуладки импорт ХИМКУРЬЕР
    chemcirier_rows_counter = []        # счетчик строк химкурьер
    date_period_of_doc = None   #ЗДЕСЬ ПРОtИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ ДЛДЯ СЕБЕСТОИМОСТИ и прайсов
    

    copy_file_created = 'Not chem courier import file'
    list_of_sheet_potential_names_var_list_is = []

    ply_dict = {}
    load_speed_index_dict = {}
    dict_of_param_to_remake_in_standart = {
            ('Сер', 'Ср'):'сер',
            ('Груз', 'Гр'):'груз',
            ('Легк', 'Лег', 'Лг'):'легк',
            ('Сх', 'С/х'):'сх',
            ('Л/г', 'л/г'):'л/г',
            ('Тр', 'Тп', 'Трп'):'трп',
            'L-2':'L-2',
            'LS-2':'LS-2',
            ('Type', 'Typ'):'Type',
            'кам':'кам',
            ('б/к', 'бк'):'б/к',
            ('ошип', 'ош', 'п/ош'):'ошип',
            'КГШ':'КГШ',
            'ЗМШ':'ЗМШ',
            'ВАЗ':'ВАЗ',
            'а/к':'а/к',
            'о/л':'о/л',
            'ак23.5':'ак23.5',
            'ол23.5':'ол23.5',
            'Газель':'Газель',
            'Вездеход':'Вездеход',
            'погр':'Погр',
            'масс':'Масс',
            ('выт', 'камневыт.'):'камневыт',
            'вен.161':'вен.161',
            'вен.ТК':'вен.ТК',
            'H':'H',
            'S':'S',
            'C':'C',
            #'РК-\d{1}-\d{3}',
            #'РК-\d{1}([А-Яа-я]|[A-Za-z])-\d{3}',
            #'ГК--\d{3}',
            #' (\d{2}|\d{1})\b',
            #'\d{3}([А-Яа-я]|[A-Za-z])\d{1}|\d{3} ([А-Яа-я]|[A-Za-z])\d{1}',
        }
    get_saved_file = None
    get_saved_file_form_a = None
    get_saved_file_form_b = None

    for f in os.listdir('media'):
        path_to = os.path.join('media',f)
        #print('RRRRRRRRRRRRRRRRRRRRRRRR', f)
        if f == 'aform_CHEM_.xlsx':
            #get_saved_file_form_a = openpyxl.load_workbook(path_to)  
            get_saved_file_form_a = openpyxl.load_workbook(path_to) 
            #print('-----+++++++---',)
            break 

    #try:
    for f in os.listdir('media'):
        path_to = os.path.join('media',f)
        #print('RRRRR000000000000000000000000RRRRRRRRR', f)
        if f == 'bform_CHEM_.xlsx': 
            #get_saved_file_form_b = openpyxl.load_workbook(path_to)  
            get_saved_file_form_b = openpyxl.load_workbook(path_to) 
            #print('-----+++++++---',)
            break 
    #except:
    #    pass

    if get_saved_file_form_a:
        try:
            #path_to_a = os.path.abspath(get_saved_file_form_a)
            #file_to_read = openpyxl.load_workbook(path_to_a, data_only=True) 
            file_to_read = get_saved_file_form_a
            #print('FILE TO READ  ===***', file_to_read )
            sheet = file_to_read['Sheet1']      # Читаем файл и лист1 книги excel
            #print(f'Total Rows = {sheet.max_row} and Total Columns = {sheet.max_column}')               # получить количество строк и колонок на листе
            # 1. Парсинг. поиск ячейки с названием 'Наименование продукции' и выборка данных из данной колонки с заголовком этой ячейки            
            for row in sheet.rows:                      
                for cell in row:
                    if cell.value == 'Наименование продукции':
                        for row in sheet.iter_rows(min_row=cell.row+1, max_row=sheet.max_row):   
                            ### проверка если строка пустая
                            if str(row[cell.column-1].value) is not str(row[cell.column-1].value):        
                                tyresize_list.append(' ')
                            #######
                            reg_list = [
                            #'\d{3}/\d{2}[A-Za-z]\d{2}(\(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2}| \(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2})', 
                            #'\d{2}(\.|\,)(\d{2}|\d{1})(R|-)\d{2}', 
                            #'(\d{3}|\d{2})/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2}[A-Za-z]|\d{2})',  # = '(\d{3}|\d{2})/\d{2}([A-Za-z]|-)\d{2}' +  '\d{3}/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2})',    
                            '(\d{3}/\d{2}[A-Za-z]\d{2}(\(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2}| \(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2}))|(\d{2}(\.|\,)(\d{2}|\d{1})(R|-)\d{2})|((\d{3}|\d{2})/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2}[A-Za-z]|\d{2}))', #3 в одном чтобы избежать повторений двойных в ячейке наподобие #АШ 480/80R42(18.4R42)
                            '\d{2}(\.|\,)\d{1}[A-Za-z](R|-)\d{2}',
                            '(\d{4}|\d{3})[A-Za-z]\d{3}([A-Za-z]|-)\d{3}',
                            '\d{2}[A-Za-z]\d{1}([A-Za-z]|-)\d{1}',
                            '\d{2}[A-Za-z]\d{1}(\.|\,)\d{2}([A-Za-z]|-)\d{1}',
                            '\d{2}(\.|\,)\d{1}/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2})',                       
                            '\d{1}(\.|\,)\d{2}(([A-Za-z]|-)|[A-Za-z]-)\d{2} ',
                            '\d{1}[A-Za-z]-\d{2}',
                            '\d{3}[A-Za-z]\d{2}[A-Za-z]',
                            '\s\d{2}([A-Za-z]|-)\d{2}(\.|\,)\d{1}', 
                            '\d{2}[A-Za-z][A-Za-z]\d{2}', 
                            ]
                            for n in reg_list:
                                result = re.search(rf'(?i){n}', str(row[cell.column-1].value))
                                if result:
                                    tyresize_list.append(result.group(0))
                                    ### удаление среза с типоразмером и всем что написано перед типоразмером
                                    left_before_size_data_index = str(row[cell.column-1].value).index(result.group(0))
                                    str_left_data = ''
                                    if left_before_size_data_index > 0:
                                        str_left_data = str(row[cell.column-1].value)[0:left_before_size_data_index-1]
                                    row[cell.column-1].value = str(row[cell.column-1].value).replace(str_left_data, '')
                                    row[cell.column-1].value = str(row[cell.column-1].value).replace(result.group(0), '')
                        for row in sheet.iter_rows(min_row=cell.row+1, max_row=sheet.max_row):    
                            ### проверка если строка пустая
                            if str(row[cell.column-1].value) is not str(row[cell.column-1].value):        
                                tyremodel_list.append(' ')
                            ####### 
                            reg_list = ['BEL-\w+',
                            '(ФБел-\d{3}-\d{1})|(Бел-\d{3}-\d{1})|ФБел-\d{3}([A-Za-z]|[А-Яа-я])|(ФБел-\d{3})|(Бел-\d{2}(\.|\,)\d{2}(\.|\,)\d{2})|(Бел-\w+)|(Бел ПТ-\w+|ПТ-\w+)|(БелОШ\w+)',
                            #'БИ-\w+',
                            #'ВИ-\w+',
                            'И-\w+|ВИ-\w+|БИ-\w+',
                            '(Ф-\d{3}-\d{1})|(Ф-\d{2}[A-Za-z][A-Za-z]-\d{1})|(Ф-\d{2}\s[A-Za-z][A-Za-z]-\d{1})|(Ф-\d{2}-\d{1})|(Ф-\w+|КФ-\w+|ВФ-\w+)',
                            'ФД-\w+',
                            'ИД-\w+',
                            '(К|K)-\d{2}[А-Яа-я][А-Яа-я]',
                            '(В-\d{2}-\d{1})|(В-\w+|ИЯВ-\w+)',
                            'ФТ-\w+',
                            'Я-\w+',]   
                            for n in reg_list:
                                result = re.search(rf'(?i){n}', str(row[cell.column-1].value))
                                if result:
                                    #print(result.group(0))
                                    tyremodel_list.append(result.group(0))
                                    ### удаление среза с моделью
                                    row[cell.column-1].value = str(row[cell.column-1].value).replace(result.group(0), '')
                                    #print(str(row[cell.column-1].value))
                        for row in sheet.iter_rows(min_row=cell.row+1, max_row=sheet.max_row):     
                            reg_list = [
                                'ЗМШ',
                                'КГШ',
                                'Сер|сер|ср',
                                'СГ',
                                'Трп|Тр|Тпр',
                                'Масс',
                                'с/х|сх',
                                'Лег|лг|легк', 
                                'Груз|груз|гр',
                                'Л/гр|Л/г|л/г',
                                'бк|б/к',
                                'Погр',
                                'кам',
                                'LS-2|LS|L-2',  
                                'Type|Typ',
                                'S|H|C',
                                '((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2})[A-Za-z]))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2}))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1})|(\d{3}|\d{2})([А-Яа-я]|[A-Za-z]))', 
                                'Газель',
                                '(ВАЗ)',
                                'Вездеход'
                                'У-\d{1}',
                                'ошип|а/к|выт|п/ош',
                                '(ГК|ЕР)-\d{3}',
                                'РК-5-165',
                                '\(ак23.5\)|\(ол23.5\)|\(ГК-260\)|вен.260|\(РК-5А-145\)|\(о/л\)|о/л|\(кам.14.9\)|\(кам12,5\)|ЛК-35-16.5|\(ГК-165\)|\(вен.ТК\)|\(вен.161\)|вен.260|Подз|\(Подз\)|вен.|(о/л)|\(ЛК-35-16.5\)',
                                '(кам.14.9)|(кам12,5)|вен.ЛК-35-16.5|ГК-145|РК.5-165',
                                '(\d{2}|\d{1})+$',
                            ] 
                            list_of_parametrs = []
                            for n in reg_list:
                                result = re.search(rf'(?i){n}', str(row[cell.column-1].value))                                  
                                if result:
                                    list_of_parametrs.append(result.group(0))   
                                    #print('rEsUlT', result, print('N is =', n))
                                    ### удаление среза с моделью                                   
                                    row[cell.column-1].value = str(row[cell.column-1].value).replace(result.group(0), '')
                                #################################### дополнительно получаем и формируем данные стандартых параметров НОРМ СЛОЙНОСТИ для добавления в словарь стандартых параметров dict_of_param_to_remake_in_standart    
                                dict_ply = ''
                                if n == '(\d{2}|\d{1})+$':
                                    if result:
                                        #print(n, 'НОРМА СЛОЙНОСТИ ПОЛУЧЕНА =', result.group(0))
                                        dict_ply = str(result.group(0))
                                        ply_dict[dict_ply] = result.group(0)
                                ###########################
                                #################################### дополнительно получаем и формируем данные индексов скорости нагрузки добавления в словарь стандартых параметров dict_of_param_to_remake_in_standart:    
                                load_speed_index = ''
                                if n == '((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2})[A-Za-z]))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2}))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1})|(\d{3}|\d{2})([А-Яа-я]|[A-Za-z]))':
                                    if result:
                                        #print(n, 'ИНДЕКС НАГРУЗКИ СКОРОСТИ ПОЛУЧЕН =', result.group(0))
                                        load_speed_index = str(result.group(0))
                                        load_speed_index_dict[load_speed_index] = result.group(0)
                                ###########################
                                pp = str(row[cell.column-1].value)                                      #### ????? это зачем - не задействовано ЖИ!
                                #print(str(row[cell.column-1].value))    
                                    ### 
                            str_of_param = ' '.join(list_of_parametrs)
                        #    print('---===-----')
                            tyreparametrs_list.append(str_of_param)
                            
            ####### очистка списка параметров шины для формирования уникальных значений для справочника модели TyreParametersModel:                    
            tyreparametrs_list_cleaned = list(set(' '.join(tyreparametrs_list).split())) 
            #print('!!!!==-=-=-', tyreparametrs_list_cleaned)
            #######################################################################################################################
            ################################# 2. модуль преобразователь - разные варианты написания к одному стандарту:
            ####### 2.1 преобразование уникальных параметров к стандарту для справочников:
            tyreparametrs_list_cleaned_and_prepared = []
            for n in (tyreparametrs_list_cleaned):
                for standart_dict_param in dict_of_param_to_remake_in_standart.keys():
                    if n in standart_dict_param:
                        n = dict_of_param_to_remake_in_standart.get(standart_dict_param)
                        tyreparametrs_list_cleaned_and_prepared.append(n)                                   # очищенные уникальные данные параметры для внесения в справочник
            ####### 2.2 преобразование параметров шины к стандарту для сопоставления с уникальными данными в справочниках в дальнейшем:
            #print(tyreparametrs_list)
            tyreparametrs_list_prepared = []
            for n in (tyreparametrs_list):
                tyre_params = ''.join(n).split()
                tyreparametrs_list_prepared.append(tyre_params)      
            ############################################################ Дополняет справочник стандартных dict_of_param_to_remake_in_standart новыми ключасми и значениями норм слойности из доп словаря ply_dict:
            dict_of_param_to_remake_in_standart.update(ply_dict)
            #print(dict_of_param_to_remake_in_standart)
            ############################################################
            ############################################################ Дополняет справочник стандартных dict_of_param_to_remake_in_standart новыми ключасми и значениями норм слойности из доп словаря load_speed_index_dict:
            dict_of_param_to_remake_in_standart.update(load_speed_index_dict)
            #print(dict_of_param_to_remake_in_standart)
            ############################################################
            #print('KK',tyreparametrs_list_prepared)       
            tyreparam_list = []    
            for tpl in tyreparametrs_list_prepared:
                tyreparametrs = []
                for n in tpl:
                    for standart_dict_param in dict_of_param_to_remake_in_standart.keys():
                        for standart_dict_param_value in standart_dict_param:
                            if n == standart_dict_param_value: 
                                #print('tyreparametr prepared=', n, 'standart_dict_param=', standart_dict_param)
                                for k in standart_dict_param:  
                                    if n == k:                                                                      #### ПРОВЕРКА НА ПОЛНОЕ соответстие с ключами dict_of_param_to_remake_in_standart для преобразования
                                        #print('n=', n, 'k=', k)
                                        n = dict_of_param_to_remake_in_standart.get(standart_dict_param)
                                        #print(n)
                                        tyreparametrs.append(n) 
                                        #print(tyreparametrs)
                    for k in list(ply_dict.values()):                                   # дополнительно провека норм слойности
                        if n == k: 
                            #print('n=',n, 'k=', k)
                            tyreparametrs.append(n)      
                    for k in list(load_speed_index_dict.values()):                                   # дополнительно провека Индексов нагрузки скорости
                        if n == k: 
                            #print('n=',n, 'k=', k)
                            tyreparametrs.append(n)         
                tyreparam_list.append(tyreparametrs)                                    # готовые данные параметры шины для сверки с справочником
            #print(tyreparametrs_list)
            #print(tyreparametrs_list_prepared)
            #print('tyreparam_list', tyreparam_list)                 
            #print(list(ply_dict.values()))
            ###################################### 3. Создать объекты справочников (ИНДЕКС СКОРОСТИ НАГРУЗИ И НОРМ СЛОЙНОСТИ ЗАКИДЫВАЮТСЯ ЗДЕСЬ ОТДЕЛЬНО):
            #for tyre_model in tyremodel_list:
            #    dictionaries_models.ModelNameModel.objects.update_or_create(model=tyre_model)
            #for tyre_size in tyresize_list:
            #    dictionaries_models.TyreSizeModel.objects.update_or_create(tyre_size=tyre_size)
            #for tyre_ply_value in ply_dict.values():
            #    dictionaries_models.TyreParametersModel.objects.update_or_create(tyre_type=tyre_ply_value)   
            #for load_speed_index_value in load_speed_index_dict.values():
            #    dictionaries_models.TyreParametersModel.objects.update_or_create(tyre_type=load_speed_index_value)   
            #for tyre_type in tyreparametrs_list_cleaned_and_prepared:
            #    dictionaries_models.TyreParametersModel.objects.update_or_create(tyre_type=tyre_type) 

            ## БОЛЕЕ БЫСТРАЯ ??????
            tyre_model_bulk_list = []
            for tyre_model in list(set(tyremodel_list)):
                tyre_model_exist = True
                try:
                    dictionaries_models.ModelNameModel.objects.get(model=tyre_model)
                except:
                    tyre_model_exist = False
                    if tyre_model_exist is False:    
                        tyre_model_bulk_list.append(dictionaries_models.ModelNameModel(model=tyre_model))
            tyre_size_bulk_list = []
            for tyre_size in list(set(tyresize_list)):
                tyre_size_exist = True
                try:
                    dictionaries_models.TyreSizeModel.objects.get(tyre_size=tyre_size)
                except:
                    tyre_size_exist = False
                    if tyre_size_exist is False:
                        tyre_size_bulk_list.append(dictionaries_models.TyreSizeModel(tyre_size=tyre_size))
            tyre_ply_value_bulk_list = []    
            for tyre_ply_value in list(set(ply_dict.values())):
                tyre_ply_value_exist = True
                try:
                    dictionaries_models.TyreParametersModel.objects.get(tyre_type=tyre_ply_value)
                except:
                    tyre_ply_value_exist = False
                    if tyre_ply_value_exist is False:
                        tyre_ply_value_bulk_list.append(dictionaries_models.TyreParametersModel(tyre_type=tyre_ply_value)) 
            load_speed_index_value_bulk_list = [] 
            for load_speed_index_value in list(set(load_speed_index_dict.values())):
                load_speed_index_value_exist = True
                try:
                    dictionaries_models.TyreParametersModel.objects.get(tyre_type=load_speed_index_value)
                except: 
                    load_speed_index_value_exist = False
                    if load_speed_index_value_exist is False:                       
                        load_speed_index_value_bulk_list.append(dictionaries_models.TyreParametersModel(tyre_type=load_speed_index_value))               
            tyre_type_value_bulk_list = [] 
            for tyre_type in list(set(tyreparametrs_list_cleaned_and_prepared)):
                tyre_type_exist = True
                try:
                    dictionaries_models.TyreParametersModel.objects.get(tyre_type=tyre_type)
                except:
                    tyre_type_exist = False
                    if tyre_type_exist is False:                         
                        tyre_type_value_bulk_list.append(dictionaries_models.TyreParametersModel(tyre_type=tyre_type))
            dictionaries_models.ModelNameModel.objects.bulk_create(tyre_model_bulk_list)
            dictionaries_models.TyreSizeModel.objects.bulk_create(tyre_size_bulk_list)
            dictionaries_models.TyreParametersModel.objects.bulk_create(tyre_ply_value_bulk_list)
            dictionaries_models.TyreParametersModel.objects.bulk_create(load_speed_index_value_bulk_list)
            dictionaries_models.TyreParametersModel.objects.bulk_create(tyre_type_value_bulk_list)  
            ## END БОЛЕЕ БЫСТРАЯ ??????                

            ##### сверка данных спарсенных с данными в справочниках:
            tyre_model_obj_list = []
            tyre_size_obj_list = []
            tyre_type_obj_list = []
            for n in dictionaries_models.ModelNameModel.objects.all():
                tyre_model_obj_list.append(n)
            for n in dictionaries_models.TyreSizeModel.objects.all():
                tyre_size_obj_list.append(n)
            for n in dictionaries_models.TyreParametersModel.objects.all():
                tyre_type_obj_list.append(n)
            tyre_model_list = [] 
            for cleaned_model_name in tyremodel_list:               # сверка спарсенных данных о модели с данными в БД справочнике модель
                for dict_ob in tyre_model_obj_list:
                    if cleaned_model_name == dict_ob.model:
                        #print(cleaned_model_name)
                        pass
                if cleaned_model_name is None:
                    cleaned_model_name = ''  
                tyre_model_list.append(cleaned_model_name)
            tyre_size_list = [] 
            for cleaned_tyresize_name in tyresize_list:               # сверка спарсенных данных о типоразмере с данными в БД справочнике типоразмер
                for dict_ob in tyre_size_obj_list:
                    if cleaned_tyresize_name == dict_ob.tyre_size:
                        #print(cleaned_tyresize_name)
                        pass
                if cleaned_tyresize_name is None:
                    cleaned_tyresize_name = ''                   
                tyre_size_list.append(cleaned_tyresize_name)
            #print('готовые данные параметры шины для сверки с справочником: tyreparam_list = ', tyreparam_list)
            tyre_param_list = []                                    # сверка спарсенных данных о параметрах шины с данными в БД справочнике параметры шины
            for cleaned_tyreparam_name in tyreparam_list:  
                trprmtrs = []   
                #print('ОТДЕЛЬНЫЙ ПОДСПИСОК парам готовые данные параметры шины для сверки с справочником', cleaned_tyreparam_name)          
                for n in cleaned_tyreparam_name:
                    #print('ОТДЕЛЬНЫЙ ЭЛЕМЕНТ ИЗ ОТДЕЛЬНЫЙ ПОДСПИСОК готовые данные параметры шины для сверки с справочником : n', n)
                    for dict_ob in tuple(tyre_type_obj_list):
                        #print('dict_ob ИЗ tyre_type_obj_list:', dict_ob.tyre_type)
                        if n == dict_ob.tyre_type:
                            #print('cleaned_tyreparam_name IS',cleaned_tyreparam_name)
                            #print('n = ', n, 'dict_ob.tyre_type = ', dict_ob.tyre_type)
                            trprmtrs.append(n)
                    if n is None:
                        n == ''
                        trprmtrs.append(n)
                tyre_param_list.append(trprmtrs)
            #print(tyre_model_list, len(tyre_model_list))
            #print(tyre_size_list, len(tyre_size_list))
            #print('tyre_param_list ==', tyre_param_list, len(tyre_param_list))
                ###### запись в модель Tyre данных о модели типоразмере и параметрах - сверенных с данными в моделях справочников
            if len(tyre_model_list) == len(tyre_size_list) and len(tyre_model_list) == len(tyre_param_list):                ########### ЗДЕСЬ ВРОДЬ КАК ДАННЫЕ ЧТО ПОЙДУТ В МОДЕЛЬ TYRE
                for n in range(0, len(tyre_param_list)):
                    ######################################  Проверяем, существует ли в БД уже объект шина с заданными параметрами:
                    mod_pos = tyre_model_list[n]
                    size_pos = tyre_size_list[n]
                    param_pos = tyre_param_list[n]
                    tyre_obj_exist = tyres_models.Tyre.objects.filter(tyre_model__model=mod_pos, tyre_size__tyre_size=size_pos, tyre_type__tyre_type__in=param_pos)
                    ######################################  
                for n in range(0, len(tyre_param_list)):
                    tyre_type_el_list = []
                    for k in range (0, len(tyre_param_list[n])):
                        tyre_type_el = dictionaries_models.TyreParametersModel.objects.get(tyre_type=tyre_param_list[n][k])
                        #print(tyre_type_el)
                        tyre_type_el_list.append(tyre_type_el)
                    if tyre_obj_exist:                              ### если объект существует - не созавать новый такой же
                        #print('YEAP!', tyre_obj_exist)
                        pass          
                    else:
                        #print('tyre_type_el_list = ', tyre_type_el_list) 
                        #### СОЗДАТЬ ОБЪЕКТ TYRES. НО! Если уже создан объект Tyres с данными параметрами то не создавать такой же:
                        tyre_obj, created = tyres_models.Tyre.objects.get_or_create(
                            tyre_model=dictionaries_models.ModelNameModel.objects.get(model=tyre_model_list[n]),
                            tyre_size=dictionaries_models.TyreSizeModel.objects.get(tyre_size=tyre_size_list[n]),
                        )
                        all_params_in_tyre = list(tyre_obj.tyre_type.all())
                        #print('all_params_in_tyre == tyre_type_el_list', all_params_in_tyre,'//////',  tyre_type_el_list)
                        if all_params_in_tyre == tyre_type_el_list:
                            pass
                        else:
                            for n in tyre_type_el_list:
                                tyre_obj.tyre_type.add(n)
                        #tyre_obj = tyres_models.Tyre.objects.create(                                                        ####  СОЗДАЕМ объект Tyre
                        #    tyre_model=dictionaries_models.ModelNameModel.objects.get(model=tyre_model_list[n]),
                        #    tyre_size=dictionaries_models.TyreSizeModel.objects.get(tyre_size=tyre_size_list[n]), 
                        #)
                        #for n in tyre_type_el_list:
                        #    tyre_obj.tyre_type.add(n)
                    #####    tyr_to_check = {"tyre_model":tyre_model_list[n], "tyre_size": tyre_size_list[n]}
                    #####    to_put_data={"tyre_type": tyre_type_el_list}
                    #####    tyres_models.Tyre.objects.update_or_create(tyr_to_check, defaults=to_put_data)
                    #################################
            ### Создать или проверить наличие групп шин:
            group_names = ['легковые', 'легкогруз', 'с/х', 'грузовые']
            for group_name in group_names:
                dictionaries_models.TyreGroupModel.objects.get_or_create(
                    tyre_group=group_name
                )
            #for obj in dictionaries_models.TyreGroupModel.objects.all():
            #    print(obj, obj.tyre_group)
            ###  
            for obj in tyres_models.Tyre.objects.all():
                for p in obj.tyre_type.all():
                    if p.tyre_type == 'груз':
                        group_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='грузовые')
                        obj.tyre_group.add(group_obect)
                    elif p.tyre_type == 'легк':
                        group_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='легковые')
                        obj.tyre_group.add(group_obect)
                    elif p.tyre_type == 'сх':
                        group_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='с/х')
                        obj.tyre_group.add(group_obect)
                    elif p.tyre_type == 'л/г':
                        group_obect = dictionaries_models.TyreGroupModel.objects.get(tyre_group='легкогруз')
                        obj.tyre_group.add(group_obect)
            ####################################################### Запись данных в файл
            from openpyxl import Workbook
            excel_file = Workbook()
            excel_sheet = excel_file.create_sheet(title='Holidays 2019', index=0)
            excel_sheet['A1'] = 'Типоразмер'
            excel_sheet['B1'] = 'Модель'
            excel_sheet['C1'] = 'Параметры'
            for n in range(0, tyres_models.Tyre.objects.all().count()):
                obj = tyres_models.Tyre.objects.all()[n]
                tyresize_val = obj.tyre_model.model
                tyremodel_val = obj.tyre_size.tyre_size
                tyreparmetrs_val_list = []
                for qo in obj.tyre_type.all():
                    tyreparmetrs_val = qo.tyre_type
                    tyreparmetrs_val_list.append(tyreparmetrs_val)
                tyreparametrs_val = ' '.join(tyreparmetrs_val_list) 
                excel_sheet.cell(row=n+1, column=1).value = tyresize_val
                excel_sheet.cell(row=n+1, column=2).value = tyremodel_val
                excel_sheet.cell(row=n+1, column=3).value = tyreparametrs_val
                #for n in range(len(tyresize_list)): 
                #    excel_sheet.cell(row=n+1, column=1).value = tyresize_list[n]
                #for n in range(len(tyremodel_list)): 
                #    excel_sheet.cell(row=n+1, column=2).value = tyremodel_list[n]
                #for n in range(len(tyreparametrs_list)): 
                #    excel_sheet.cell(row=n+1, column=3).value = tyreparametrs_list[n]     
            #for n in range(len(pp)): 
            #    excel_sheet.cell(row=n+1, column=4).value = pp[n]                         
            excel_file.save(filename="Tyres.xlsx")
    ############################################################
            form = forms.ImportDataForm()
            file_to_read.close()
        #    return render(self.request, 'filemanagment/excel_import.html', {'form': form})        
    ##########################################################################   ##### ПАРСИНГ ХИМКУРЬЕРА #########
    #### ЕСЛИ ЗАБРАСЫВАЮТСЯ ФАЙЛЫ ХИМКУРЬЕР:

        except:
            pass   
  
    if get_saved_file_form_b:  
    #try:
        flag_chem_import = False
        list_of_sheet_potential_names_var_list = ['ИМПОРТ', 'импорт',]
        sheet = None
        #path_to_b = os.path.abspath(get_saved_file_form_b)
        #file_to_read = openpyxl.load_workbook(path_to_b, data_only=True)


        file_to_read = get_saved_file_form_b
        list_chemcurier_years = range(11, 50)
        for year_d in list_chemcurier_years:
            for sheet_name in list_of_sheet_potential_names_var_list:
                try:
                    sheet = file_to_read[f'{sheet_name} 20{year_d}']
                    #sheet = file_to_read['ИМПОРТ 2022']    # Читаем файл и лист1 книги excel 
                except:
                    pass
        #print('SSSHHHEEEEt', sheet)            
        if sheet:
            ############ создать физическую копию файла юзера для работы с ней 
            # необходим для работы "за кадром" вне POST запроса
            # opening the source excel file 
            # если есть файл от юзера для хим курьера - работаем дальше
            if file_to_read:
                #print('PPPPOOOGNALYYY')
                flag_chem_import = True 
                file_to_copy = file_to_read
                sheet_to_copy = sheet
                copy_file_created = make_a_copy_of_users_chemc_file(file_to_copy, sheet_to_copy)
                list_of_sheet_potential_names_var_list_is = list_of_sheet_potential_names_var_list
                rows_in_file_limiter(copy_file_created, list_of_sheet_potential_names_var_list)
                file_to_read.close()
                delete_temp_file()
                return copy_file_created, list_of_sheet_potential_names_var_list_is
            # если нет файла от юзера для хим курьера - скипнуть
            else:
                file_to_read.close()
                return copy_file_created,  list_of_sheet_potential_names_var_list_is 
    ##print('tiresize_chemcurier_dict', tiresize_chemcurier_dict)
    ##print('tire_group_chemcurier_dict', tire_group_chemcurier_dict)
    #print('month_chemcurier_dict', pieces_month_chemcurier_dict)
    #print('money_month_chemcurier_dic', money_month_chemcurier_dict)
    #print('date_pieces_row_chemcurier_dict', date_pieces_row_chemcurier_dict)
    #print('recipipient_chemcurier_dict', recipipient_chemcurier_dict)
    #print('prod_country_column_dict', prod_country_column_dict)
    ##################### END ХИМКУРЬЕР Ч.1
######################################################################### END ХИМКУРЬЕР ПАРРСИНГА
#########################################################################
### ЕСЛИ ЗАБРАСЫВАЮТСЯ ФАЙЛЫ С ДАННЫМИ О ПРОДАЖАХ/ОСТАТКАХ/ПРОИЗВОДСТВЕ/ЦЕНЫ:
################ считывание файла и сопоставление с текущей базой данных
#elif form_name == "bform.prefix":
#    form = forms.ImportSalesDataForm(self.request.POST, self.request.FILES)  
#    if form.is_valid():
#        file_to_read = openpyxl.load_workbook(self.request.FILES['file_fields'], data_only=True)  
        else:
            if flag_chem_import is False:
                try:   
                    sheet = file_to_read['Sheet1']      # Читаем файл и лист1 книги excel 
                    for row in sheet.rows:                    
                        for cell in row:      
                           # 1 Парсинг    
                            # ПАРСИНГ ДАТЫ ДЛЯ ТАБЛИЦЫ ДАННЫХ МИНИМАЛКИ И ПРОЧЕЕ
                            if isinstance(cell.value, datetime):
                                date_period_of_doc = cell.value.date()
                            #    print('cell.is_date === !!!!!!!!!!!!!!!!!!!!!!!!!!!', date_period_of_doc)        
                            ###############   
                            #                      
                            if cell.value == 'контрагент':          # получаем колонку 'контрагент'
                                contragent_column = cell.column
                                contragent_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=contragent_row+1, min_col=contragent_column, max_col=contragent_column, max_row=sheet.max_row):
                                        for cell in col:
                                            contragent_value = ''
                                            contragent_value =  cell.value                               
                                            contragent_list.append(contragent_value)
                                except:
                                    pass
                            elif cell.value == 'объем продаж':
                                saless_row_temp = int
                                #sales_column = cell.coordinate          # получаем колонку 'объем продаж'
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value                            
                                            sales_list.append(sell_value)
                                            #print('saless_row', saless_row, 'current_row_number', current_row_numberr, cell.row)
                                            saless_row_dict[cell.row] = sell_value                                                      # закидываем в словарь строка значение объем продаж
                                            pass  
                                except:
                                    pass     
                                sales_list = [float(x) for x in sales_list]    # str значения в float
                            elif cell.value == 'Полные затраты':
                                #sales_column = cell.coordinate          # получаем колонку 'полные затраты'    planned_costs
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                #print('cell.value.lstrip().rstrip()', cell.value.lstrip().rstrip() )
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value  
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                                       
                                                planned_costs.append(sell_value)
                                                planned_costs_row_dict[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                #print(planned_costs)
                                planned_costs = [float(x) for x in planned_costs]    # str значения в float
                            elif cell.value == 'прямые затраты':
                                #sales_column = cell.coordinate          # получаем колонку 'прямые затраты'    semi_variable_costs
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value   
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                          
                                                semi_variable_costs.append(sell_value)
                                                semi_variable_costs_row_dict[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                #print(semi_variable_costs)
                                semi_variable_costs = [float(x) for x in semi_variable_costs]    # str значения в float
                                #print(semi_variable_costs)
                            elif cell.value == 'прейскуранты №№9, 902':
                                #sales_column = cell.coordinate          # получаем колонку 'прейскуранты №№9, 902'    belarus902price_costs
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                              
                                                belarus902price_costs.append(sell_value)
                                                belarus902price_costs_row_dict[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                belarus902price_costs = [float(x) for x in belarus902price_costs]    # str значения в float
                                #print(belarus902price_costs)
                            elif cell.value == 'ТПС РФ FCA':
                                #sales_column = cell.coordinate          # получаем колонку 'ТПС РФ FCA'    tpsrussiafcaprice
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                              
                                                tpsrussiafcaprice_costs.append(sell_value)
                                                tpsrussiafcaprice_costs_row[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                tpsrussiafcaprice_costs = [float(x) for x in tpsrussiafcaprice_costs]    # str значения в float
                                #print(tpsrussiafcaprice_costs)
                            elif cell.value == 'ТПС Казахстан FCA':
                                #sales_column = cell.coordinate          # получаем колонку 'ТПС Казахстан FCA'    tpskazfcaprice
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                              
                                                tpskazfcaprice_costs.append(sell_value)
                                                tpskazfcaprice_cost_row[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                tpskazfcaprice_costs = [float(x) for x in tpskazfcaprice_costs]    # str значения в float
                                #print('tpskazfcaprice_costs', tpskazfcaprice_costs)
                            elif cell.value == 'ТПС Средняя Азия, Закавказье, Молдова FCA':
                                #sales_column = cell.coordinate          # получаем колонку 'ТПС Средняя Азия, Закавказье, Молдова FCA'    tpsmiddleasiafcaprice
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                              
                                                tpsmiddleasiafcaprice_costs.append(sell_value)
                                                tpsmiddleasiafcaprice_costs_row[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                tpsmiddleasiafcaprice_costs = [float(x) for x in tpsmiddleasiafcaprice_costs]    # str значения в float
                                #print(tpsmiddleasiafcaprice_costs)
                            elif cell.value == 'Действующие цены':
                                #sales_column = cell.coordinate          # получаем колонку 'Действующие цены'   
                                sales_column = cell.column
                                sales_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=sales_row+1, min_col=sales_column, max_col=sales_column, max_row=sheet.max_row):
                                        for cell in col:
                                            sell_value = ''
                                            #sell_value =  cell.value.lstrip().rstrip().replace('.', ',')      # убрать пробелы в начале строки и в конце строки  
                                            if sell_value is str:
                                                sell_value = cell.value.lstrip().rstrip()   
                                            else:
                                                sell_value = cell.value
                                            if sell_value is None:
                                                pass
                                            else:
                                                if type(sell_value) is str:
                                                    sell_value = 0                              
                                                current_prices.append(sell_value)
                                                current_prices_row[cell.row] = sell_value                                                      # закидываем в словарь строка значение 
                                            pass
                                except:
                                    pass
                                current_prices = [float(x) for x in current_prices]    # str значения в float
                                #print(current_prices)
                                #print('current_prices_row', current_prices_row)
                            elif cell.value == 'индексы':          # получаем колонку 'индексы'   ДОПОЛНИТЕЛЬНЫЕ
                                indexes_column = cell.column
                                indexes_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=indexes_row+1, min_col=indexes_column, max_col=indexes_column, max_row=sheet.max_row):
                                        for cell in col:
                                            indexes_value = ''
                                            indexes_value =  cell.value                               
                                            indexes_row_dict[cell.row] = cell.value
                                except:
                                    pass
                            #print('indexes_list', indexes_list)
                            #print('indexes_row_dict', indexes_row_dict)
                            elif cell.value == 'сезонность':          # получаем колонку 'сезонность' ДОПОЛНИТЕЛЬНЫЕ
                                season_column = cell.column
                                season_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=season_row+1, min_col=season_column, max_col=season_column, max_row=sheet.max_row):
                                        for cell in col:
                                            season_value = ''
                                            season_value =  cell.value                               
                                            season_row_dict[cell.row] = cell.value
                                except:
                                    pass
                            #print('season_row_dict', season_row_dict)
                            elif cell.value == 'рисунок протектора':          # получаем колонку 'рисунок протектора' ДОПОЛНИТЕЛЬНЫЕ
                                thread_column = cell.column
                                thread_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=thread_row+1, min_col=thread_column, max_col=thread_column, max_row=sheet.max_row):
                                        for cell in col:
                                            threadn_value = ''
                                            thread_value =  cell.value                               
                                            thread_row_dict[cell.row] = cell.value
                                except:
                                    pass
                            #print('thread_row_dict', thread_row_dict)
                            elif cell.value == 'ось':          # получаем колонку 'ось' ДОПОЛНИТЕЛЬНЫЕ
                                ax_column = cell.column
                                ax_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=ax_row+1, min_col=ax_column, max_col=ax_column, max_row=sheet.max_row):
                                        for cell in col:
                                            ax_value = ''
                                            ax_value =  cell.value                               
                                            ax_row_dict[cell.row] = cell.value 
                                except:
                                    pass
                            #print('ax_row_dict', ax_row_dict)
                            elif cell.value == 'применяемость':          # получаем колонку 'применяемость' ДОПОЛНИТЕЛЬНЫЕ
                                usability_column = cell.column
                                usability_row = cell.row
                                try:
                                    for col in sheet.iter_cols(min_row=usability_row+1, min_col=usability_column, max_col=usability_column, max_row=sheet.max_row):
                                        for cell in col:
                                            usability_value = ''
                                            usability_value =  cell.value                               
                                            usability_row_dict[cell.row] = cell.value
                                except:
                                    pass
                            #print('usability_row_dict', usability_row_dict)
                            #if cell.value == 'дата':        # получаем строку'дата'
                            #    #print(cell.value)    
                            #    #print(cell.coordinate) 
                            #    cell = sheet.cell(row=cell.row+1, column=cell.column)
                            #    column_sell_date = cell.value
                            #    date_period = column_sell_date                      # ЗДЕСЬ ПОЛУЧЕНА ДАТА ДЛЯ СЕБЕСТОИМОСТИ И ПРАЙСОВ
                            ##elif cell.is_date:          # получаем дату, для работы с периодом действия цен минималок в дальнейшем
                            ##    date_period_of_doc = cell.value.date()
                            ##    ##print('cell.is_date ===', cell.is_date, 'cell.is_date ===', cell.value, 'cell.is_date ===', cell.value.date())
                            ##    print('date_period ======= ', date_period_of_doc)
                            ##  ПОЛУЧЕНИЕ МОДЕЛИ ТИПОРАЗМЕРА и ТИПА ДЛЯ ФОРМИРОВАНИЯ СЛОВАРЯ И СВЕРКИ СОСПАВШИХ ШИН ИЗ БД ДЛЯ ВЫБОРКИ ДАННЫХ ПРДАЖИ И МИНИМАЛКИ И ПР ИЗ ЭТОЙ СТРОКИ  !!!!!!!!!!!
                            ##
                            current_row_number = int
                            if cell.value == 'Наименование продукции':
                                for row in sheet.iter_rows(min_row=cell.row+1, max_row=sheet.max_row):   
                                    print('Row number:', str(row[0].row),'ROW ROW ROW')                         # СПОСОБ ПОЛУЧИТЬ НОМЕР СТРОКИ
                                    ### проверка если строка пустая
                                    if str(row[cell.column-1].value) is not str(row[cell.column-1].value):        
                                        tyresize_list.append(' ')
                                    #######
                                    reg_list = [
                                    #'\d{3}/\d{2}[A-Za-z]\d{2}(\(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2}| \(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2})', 
                                    #'\d{2}(\.|\,)(\d{2}|\d{1})(R|-)\d{2}', 
                                    #'(\d{3}|\d{2})/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2}[A-Za-z]|\d{2})',  # = '(\d{3}|\d{2})/\d{2}([A-Za-z]|-)\d{2}' +  '\d{3}/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2})',    
                                    '(\d{3}/\d{2}[A-Za-z]\d{2}(\(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2}| \(\d{2}(\.|\,)\d{1}[A-Za-z]\d{2}))|(\d{2}(\.|\,)(\d{2}|\d{1})(R|-)\d{2})|((\d{3}|\d{2})/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2}[A-Za-z]|\d{2}))', #3 в одном чтобы избежать повторений двойных в ячейке наподобие #АШ 480/80R42(18.4R42)
                                    '\d{2}(\.|\,)\d{1}[A-Za-z](R|-)\d{2}',
                                    '(\d{4}|\d{3})[A-Za-z]\d{3}([A-Za-z]|-)\d{3}',
                                    '\d{2}[A-Za-z]\d{1}([A-Za-z]|-)\d{1}',
                                    '\d{2}[A-Za-z]\d{1}(\.|\,)\d{2}([A-Za-z]|-)\d{1}',
                                    '\d{2}(\.|\,)\d{1}/\d{2}([A-Za-z]|-)(\d{2}(\.|\,)\d{1}|\d{2})',                       
                                    '\d{1}(\.|\,)\d{2}(([A-Za-z]|-)|[A-Za-z]-)\d{2} ',
                                    '\d{1}[A-Za-z]-\d{2} ',
                                    '\d{3}[A-Za-z]\d{2}[A-Za-z]',
                                    '\s\d{2}([A-Za-z]|-)\d{2}(\.|\,)\d{1}', 
                                    '\d{2}[A-Za-z][A-Za-z]\d{2}', 
                                    ]
                                    for n in reg_list:
                                        result = re.search(rf'(?i){n}', str(row[cell.column-1].value))
                                        if result:
                                            #print('result.group(0)', result.group(0))
                                            tyresize_list.append(result.group(0))
                                            tyretype_row_dict[row[0].row] = result.group(0)                                     # закидываем в словарь строка значение 
                                            #print(row[0].row, 'cell.row', result.group(0), 'result.group(0)')
                                            ### удаление среза с типоразмером и всем что написано перед типоразмером
                                            left_before_size_data_index = str(row[cell.column-1].value).index(result.group(0))
                                            if left_before_size_data_index > 0:
                                                str_left_data = str(row[cell.column-1].value)[0:left_before_size_data_index-1]
                                            row[cell.column-1].value = str(row[cell.column-1].value).replace(str_left_data, '')
                                            row[cell.column-1].value = str(row[cell.column-1].value).replace(result.group(0), '')
                                for row in sheet.iter_rows(min_row=cell.row+1, max_row=sheet.max_row):    
                                    ### проверка если строка пустая
                                    if str(row[cell.column-1].value) is not str(row[cell.column-1].value):        
                                        tyremodel_list.append(' ')
                                    ####### 
                                    reg_list = ['BEL-\w+',
                                    '(ФБел-\d{3}-\d{1})|(Бел-\d{3}-\d{1})|ФБел-\d{3}([A-Za-z]|[А-Яа-я])|(ФБел-\d{3})|(Бел-\d{2}(\.|\,)\d{2}(\.|\,)\d{2})|(Бел-\w+)|(Бел ПТ-\w+|ПТ-\w+)|(БелОШ\w+)',
                                    #'БИ-\w+',
                                    #'ВИ-\w+',
                                    'И-\w+|ВИ-\w+|БИ-\w+',
                                    '(Ф-\d{3}-\d{1})|(Ф-\d{2}[A-Za-z][A-Za-z]-\d{1})|(Ф-\d{2}\s[A-Za-z][A-Za-z]-\d{1})|(Ф-\d{2}-\d{1})|(Ф-\w+|КФ-\w+|ВФ-\w+)',
                                    'ФД-\w+',
                                    'ИД-\w+',
                                    '(К|K)-\d{2}[А-Яа-я][А-Яа-я]',
                                    '(В-\d{2}-\d{1})|(В-\w+|ИЯВ-\w+)',
                                    'ФТ-\w+',
                                    'Я-\w+',]   
                                    for n in reg_list:
                                        result = re.search(rf'(?i){n}', str(row[cell.column-1].value))
                                        if result:
                                            #print(result.group(0), 'result.group(0)')
                                            tyremodel_list.append(result.group(0))
                                            model_row_dict[row[0].row] = result.group(0)        # закидываем в словарь строка значение 
                                            model_row = result.group(0)
                                            ### удаление среза с моделью
                                            row[cell.column-1].value = str(row[cell.column-1].value).replace(result.group(0), '')
                                            #print(str(row[cell.column-1].value))
                                for row in sheet.iter_rows(min_row=cell.row+1, max_row=sheet.max_row):     
                                    reg_list = [
                                        'ЗМШ',
                                        'КГШ',
                                        'Сер|сер|ср',
                                        'СГ',
                                        'Трп|Тр|Тпр',
                                        'Масс',
                                        'с/х|сх',
                                        'Лег|лг|легк', 
                                        'Груз|груз|гр',
                                        'Л/гр|Л/г|л/г',
                                        'бк|б/к',
                                        'Погр',
                                        'кам',
                                        'LS-2|LS|L-2',  
                                        'Type|Typ',
                                        'S|H|C',
                                        '((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2})[A-Za-z]))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2}))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1})|(\d{3}|\d{2})([А-Яа-я]|[A-Za-z]))', 
                                        'Газель',
                                        '(ВАЗ)',
                                        'Вездеход'
                                        'У-\d{1}',
                                        'ошип|а/к|выт|п/ош',
                                        '(ГК|ЕР)-\d{3}',
                                        'РК-5-165',
                                        '\(ак23.5\)|\(ол23.5\)|\(ГК-260\)|вен.260|\(РК-5А-145\)|\(о/л\)|о/л|\(кам.14.9\)|\(кам12,5\)|ЛК-35-16.5|\(ГК-165\)|\(вен.ТК\)|\(вен.161\)|вен.260|Подз|\(Подз\)|вен.|(о/л)|\(ЛК-35-16.5\)',
                                        '(кам.14.9)|(кам12,5)|вен.ЛК-35-16.5|ГК-145|РК.5-165',
                                        '(\d{2}|\d{1})+$',
                                    ]                
                                    list_of_parametrs = []
                                    for n in reg_list:
                                        result = re.search(rf'(?i){n}', str(row[cell.column-1].value)) 
                                        if result:
                                            list_of_parametrs.append(result.group(0)) 
                                            #print(row[0].row, result.group(0), type(result.group(0)))
                                            params_row_dict[row[0].row] = list_of_parametrs           # закидываем в словарь строка значение 
                                            #print('rEsUlT', result, print('N is =', n))
                                            ### удаление среза с моделью                                   
                                            row[cell.column-1].value = str(row[cell.column-1].value).replace(result.group(0), '')
                                        #################################### дополнительно получаем и формируем данные стандартых параметров НОРМ СЛОЙНОСТИ для добавления в словарь стандартых параметров dict_of_param_to_remake_in_standart    
                                        dict_ply = ''
                                        if n == '(\d{2}|\d{1})+$':
                                            if result:
                                                #print(n, 'НОРМА СЛОЙНОСТИ ПОЛУЧЕНА =', result.group(0))
                                                dict_ply = str(result.group(0))
                                                ply_dict[dict_ply] = result.group(0)
                                        ###########################
                                        #################################### дополнительно получаем и формируем данные индексов скорости нагрузки добавления в словарь стандартых параметров dict_of_param_to_remake_in_standart:    
                                        load_speed_index = ''
                                        if n == '((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2})[A-Za-z]))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1}((\d{3}|\d{2}))|((\d{3}|\d{2})([А-Яа-я]|[A-Za-z])\d{1})|(\d{3}|\d{2})([А-Яа-я]|[A-Za-z]))':
                                            if result:
                                                #print(n, 'ИНДЕКС НАГРУЗКИ СКОРОСТИ ПОЛУЧЕН =', result.group(0))
                                                load_speed_index = str(result.group(0))
                                                load_speed_index_dict[load_speed_index] = result.group(0)
                                        ###########################
                                        pp = str(row[cell.column-1].value)                                      #### ????? это зачем - не задействовано ЖИ!
                                        #print(str(row[cell.column-1].value))    
                                            ### 
                                    str_of_param = ' '.join(list_of_parametrs)
                                    tyreparametrs_list.append(str_of_param)
                                    #print(str_of_param, 'str_of_param')
                    ############################################################ Дополняет справочник стандартных dict_of_param_to_remake_in_standart новыми ключасми и значениями норм слойности из доп словаря ply_dict:
                    dict_of_param_to_remake_in_standart.update(ply_dict)
                    #print(dict_of_param_to_remake_in_standart)
                    ############################################################
                    ############################################################ Дополняет справочник стандартных dict_of_param_to_remake_in_standart новыми ключасми и значениями норм слойности из доп словаря load_speed_index_dict:
                    dict_of_param_to_remake_in_standart.update(load_speed_index_dict)
                    #print(dict_of_param_to_remake_in_standart)
                    ############################################################
                    # ПРЕОБРАЗОВАНИЕ СПАРСЕННЫХ ДАННЫХ В ВИД ПО БД   + при этом далее есть еще старый код по преобразованию не словаря а списка всех значений
                    for keys, values in params_row_dict.items():
                        #print(keys, values)
                        tyreparametrs_list_clean = []
                        for n in (values):
                            tyre_el = sorted(n.split(' '), reverse=True)
                            tyreparametrs_list_cleaned_and_prepared = []
                            for el_in_param in tyre_el:
                                for standart_dict_param in dict_of_param_to_remake_in_standart.keys():
                                    #print('HHHHHHHHH', type(tuple(standart_dict_param)), standart_dict_param)           ###### str значения словаря в тапл
                                    if el_in_param in list(standart_dict_param) or el_in_param == standart_dict_param: 
                                        #print('el_in_param = ', el_in_param, 'standart_dict_param = ', standart_dict_param)      #!!!!!!!!!!!!!!!!
                                        el_in_param = dict_of_param_to_remake_in_standart.get(standart_dict_param)
                                        tyreparametrs_list_cleaned_and_prepared.append(el_in_param)  
                                el = ' '.join(tyreparametrs_list_cleaned_and_prepared) 
                            tyreparametrs_list_clean.append(el)  
                        params_row_dict[keys] = tyreparametrs_list_clean
                except:
                    pass
        
        #except:
        #    pass
            
                        
            

        
        #### СОЗДАТЬ ОБЪЕКТЫ ХИМКУРЬЕР: !!!!!!!!!!!!!!!!!!!!!!!!!!!!

        #chem_courier_bulk_write_ib_bd(MAIN_chemcirier_import_dict)

        # END BULK_CREATE  СОЗДАТЬ ОБЪЕКТЫ ХИМКУРЬЕР !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    #        che_curier_obj_tyre = che_curier_obj_tyre
        # 1.3 создать объекты продажа на дату:   
    #        dates_range = len(val[2])
    #        list_of_pr_objs = []
    #        for n in range(0, dates_range):
    #            #print(val[2][n], val[3][n])
    #            date_month = val[2][n][0]
    #            val_month = val[2][n][1]
    #            money_month = val[3][n][1]
    #            if money_month and val_month:
    #                price_on_date_chem = money_month / val_month
    #            else:
    #                price_on_date_chem = None
    #            print(date_month, val_month, money_month, money_month)  
    #            pr_obj = prices_models.DataPriceValMoneyChemCurierModel.objects.update_or_create(
    #                data_month_chem = date_month,
    #                val_on_moth_chem = val_month,
    #                money_on_moth_chem = money_month,  
    #                price_on_date_chem = price_on_date_chem,
    #                price_val_money_data = che_curier_obj_tyre,
    #            )
    #            list_of_pr_objs.append(pr_obj)
    #        print('list_of_pr_objs', list_of_pr_objs)
    #            
    #    all_checur_objs = prices_models.ChemCurierTyresModel.objects.all()
    #    #for obj in all_checur_objs:
    #    #    print('!!!!', obj, obj.price_val_money_data_obj.all())'
    #    ########################### END ХИКУРЬЕР Ч.2
        
        ############################################################ Дополняет справочник стандартных dict_of_param_to_remake_in_standart новыми ключасми и значениями норм слойности из доп словаря ply_dict:
        dict_of_param_to_remake_in_standart.update(ply_dict)
        #print(dict_of_param_to_remake_in_standart)
        ############################################################
        ############################################################ Дополняет справочник стандартных dict_of_param_to_remake_in_standart новыми ключасми и значениями норм слойности из доп словаря load_speed_index_dict:
        dict_of_param_to_remake_in_standart.update(load_speed_index_dict)
        #print(dict_of_param_to_remake_in_standart)
        ############################################################
        ####### 2. преобразование параметров шины к стандарту для сопоставления с уникальными данными в справочниках в дальнейшем:
        tyreparametrs_list_cleaned_and_prepared_ready = []                  # список параметов шины из таблицы подготовленных для сопоставления с базой
        for n in (tyreparametrs_list):
            tyre_el = sorted(n.split(' '), reverse=True)
            tyreparametrs_list_cleaned_and_prepared = []
            for el_in_param in tyre_el:
                for standart_dict_param in dict_of_param_to_remake_in_standart.keys():
                    #print('HHHHHHHHH', type(tuple(standart_dict_param)), standart_dict_param)           ###### str значения словаря в тапл
                    if el_in_param in list(standart_dict_param) or el_in_param == standart_dict_param: 
                        #print('el_in_param = ', el_in_param, 'standart_dict_param = ', standart_dict_param)      #!!!!!!!!!!!!!!!!
                        el_in_param = dict_of_param_to_remake_in_standart.get(standart_dict_param)
                        tyreparametrs_list_cleaned_and_prepared.append(el_in_param)  
                el = ' '.join(tyreparametrs_list_cleaned_and_prepared)                   
            tyreparametrs_list_cleaned_and_prepared_ready.append(el)  # список параметов шины из таблицы подготовленных для сопоставления с базой
            
        tyre_param_list_from_db = dictionaries_models.TyreParametersModel.objects.all()
        tyre_param_list_from_db_param_only = []
        for n in tyre_param_list_from_db:
            tyre_param_list_from_db_param_only.append(n.tyre_type)              #####! спикок уникальных имет параметров шины из базы данных
        #print('данные из БД перечень:', tyre_param_list_from_db_param_only)      
        #print('исходные данные:', tyreparametrs_list)             
        #print('подготовленные для сверки с бд:', tyreparametrs_list_cleaned_and_prepared_ready) 
    
        ######### 3 Сопоставление - сверка данных из таблицы с параметами в БД
        
        #print(tyreparametrs_list_cleaned_and_prepared_ready)
        #print(tyre_param_list_from_db_param_only)
        tyreparam_list = []                                                     # ПОДГОТОВЛЕНЫЫЙ И ОЧИЩЕННЫЙ СПИСОК ДАННЫХ ПАРАМЕТРОВ ШИНЫ ДЛЯ СОПОСТАВЛЕНИЯ С БД НА ФИНАЛЬНОЙ СТАДИИ ВНЕСЕНИЯ (реализац/остатки/произв) в Базу  
        tyreparam_list_with_lists = []                                             # ПОДГОТОВЛЕНЫЫЙ И ОЧИЩЕННЫЙ СПИСОК ДАННЫХ ПАРАМЕТРОВ ШИНЫ в виде вложенных списков ДЛЯ СОПОСТАВЛЕНИЯ С БД НА ФИНАЛЬНОЙ СТАДИИ ВНЕСЕНИЯ (реализац/остатки/произв) в Базу  
        for tpl in tyreparametrs_list_cleaned_and_prepared_ready:
            tyre_params_value = []
            tpl_list = sorted(tpl.split(' '), reverse=True)
            for n in tpl_list:  # sorted(GG.split(' '), reverse=True)   
                for standart_dict_param in tyre_param_list_from_db_param_only:
                    if n == standart_dict_param:
                    #if n in standart_dict_param:
                        tyre_params_value.append(n)
            tyre_params_value_str = ', '.join(tyre_params_value)                    ## !!!! Для формирования списка со строками 
            tyreparam_list.append(tyre_params_value_str)  
            tyreparam_list_with_lists.append(tyre_params_value) 
        #print(len(tyreparam_list), tyreparam_list)                               # совпавшие данные параметры шины из таблицы с параметрами в БД
        #print(len(tyresize_list), tyresize_list)
        #print(len(tyremodel_list), tyremodel_list)
        #print(len(tyreparam_list_with_lists), tyreparam_list_with_lists)
        ##### Запись в модель данных (например, о продажах):
        #### Проверка - вывод в файл данных
        #if len(tyresize_list) == len(tyremodel_list) and len(tyresize_list) ==  len(tyreparametrs_list):
        #    for n in range(0, len(tyresize_list)):
        #        #if tyresize_list[n] == tyres_models.Tyre.objects.filter(tyre_size__tyre_size=tyresize_list[n]) and tyremodel_list[n] == tyres_models.Tyre.objects.filter(tyre_model__model=tyremodel_list[n]) and tyreparametrs_list[n] == tyres_models.Tyre.objects.filter(tyre_type__tyre_type=tyreparametrs_list[n]):
        #        some_list = tyres_models.Tyre.objects.filter(tyre_size__tyre_size__icontains='315/80R22.5')
        #        print(some_list, tyresize_list[n], tyres_models.Tyre.objects.get(id=2).tyre_size)
        ###################################### СВЕРКА СПИСКОВ МОДЕЛЬ ТИПОРАЗМЕР ПАРАМЕТРЫ с ДАННЫМИ В МОДЕЛИ ШИНЫ В БД  + внесение данныхв в одель Sales(Например):
        #print(len(tyresize_list), tyresize_list)           # 1   
        #print(len(tyremodel_list), tyremodel_list)         # 2
        #print(len(tyremodel_list), tyreparam_list)          # 3/1  в виде списка со строками
        ##print(len(tyreparam_list_with_lists), tyreparam_list_with_lists)    # 3/2  в виде списка с вложенными списками

        #print(len(tyresize_list), 'tyresize_list')           # 1   
        #print(len(tyremodel_list), 'tyremodel_list')         # 2
        #print(len(tyremodel_list), 'tyreparam_list')          # 3/1  в виде списка со строками
        #print('SALES', len(sales_list), sales_list)
        #print(tyres_models.Tyre.objects.get(tyre_type__tyre_type__contains='сер'))
        #print(tyres_models.Tyre.objects.filter(tyre_type__tyre_type__in=tyreparam_list_with_lists[2]))
        tyre_obj_list = []
        sale_obj_list = []
        object_list_finel = []
        #### прокладка: уберем запятые из строк с параметрами:
        #print('TTT', tyreparam_list)
        t_list = []
        for n in tyreparam_list:
            temporaly_list = []
            param = str(n)
            n = param.replace(',', '')
            temporaly_list = n
            t_list.append(temporaly_list)
        tyreparam_list = t_list
        #print('TTT',tyreparam_list)
        ####
        ######################################                      ВАРИАНТ  1  =  СОПОСТАВЛЕНИЕ СОВПАШИХ ШИН С БД
        list_of_unique_tyres_objs_cleaned = []
        #print(tyreparam_list)
        if len(tyresize_list) == len(tyremodel_list) and len(tyremodel_list) == len(tyreparam_list):
            for n in range(0, len(tyreparam_list)):
                ######################################  Проверяем, существует ли в БД уже объект шина с заданными параметрами:
                mod_pos = tyremodel_list[n]
                size_pos = tyresize_list[n]
                param_pos = tyreparam_list[n]
                #print(mod_pos, size_pos, param_pos)
                param_pos_list = sorted(tyreparam_list[n].split(' '), reverse=True)
                #print(param_pos_list)
                tyre_obj_exist = tyres_models.Tyre.objects.filter(tyre_model__model=mod_pos, tyre_size__tyre_size=size_pos, tyre_type__tyre_type__in=param_pos_list)
                if tyre_obj_exist:                              ### если кверисет объектов существует
                    #print('YEAP!', tyre_obj_exist)
                #print('_____________')
                    list_of_unique_tyres_objs = []
                    for obj in tyre_obj_exist:              ### берем объекты из кверисета
                        #print('param=',param, 'obj.tyre_type.all() = ', obj.tyre_type.all())
                        counter_coinc = 0
                        for param in param_pos_list:                ### берем параметры из param_pos_list и перебираем
                            len(param_pos_list)                     ###  замеряем количество параметров в текущем param_pos_list 
                            #print(len(param_pos_list))
                            for p in obj.tyre_type.all():       ###   перебираем параметры из param_pos_list в параметрах объекта
                                #print('param = ', param , 'p.tyre_type = ', p.tyre_type)
                                if param == p.tyre_type:        ### если параметры совпадают
                                    #print('param = ', param, 'p.tyre_type = ', p.tyre_type)
                                    counter_coinc += 1
                        if counter_coinc == len(param_pos_list):        ###### ЗДЕСЬ ПОЛУЧАЕМ ОБЪЕКТ У КОТОРОГО ВСЕ ПАРАМЕТРЫ СОВПАЛИ СПАРАМЕТРАМИ ШИНЫ В БАЗЕ
                            ##################print(obj)
                            list_of_unique_tyres_objs.append(obj)
                    list_of_unique_tyres_objs_cleaned.append(list(set(list_of_unique_tyres_objs)))
                #print('_____________')
        #print(list_of_unique_tyres_objs_cleaned, len(list_of_unique_tyres_objs_cleaned))        ####!!!!!! СПИСОК С ОБЪЕКТАМИ с СОВПАВШИМИ ПАРАМЕТРАМИ - итоговые даннее готовые
        #### ВСТАВКА:
        ##### + Создаем единую таблицу SALES_TABLE:                                                
        sales_table, created = sales_models.SalesTable.objects.get_or_create()                   ##### !!!!!  ####### ДЛЯ РАБОТЫ С СТРАНИЦЕЙ SALES:    
        #### КОНЕЦ ВСТАВКИ  
        ### ЗДЕСЬ ПОНАДОБИТСЯ ДАЛЬШЕ когда получим нужный объект:
        #print(list_of_unique_tyres_objs_cleaned)
        sales_list_values_list = sales_list
        obj_list_el_bulk_list = [] 
        for obj_list_el in list_of_unique_tyres_objs_cleaned:
            row_value_counter = 0
            #print(obj_list_el)
            # Дата продаж
            if column_sell_date:
                #print(type(column_sell_date))
                #date_time_obj = datetime.strptime(column_sell_date, '%Y/%m/%d')
                #sell_date = date_time_obj
                sell_date = column_sell_date
            else:
                sell_date =  datetime.today()
            # Объемы продаж
            if sales_list_values_list:
                n = sales_list_values_list.pop(0)
            else:
                n = 2
            # Контрагенты:
            #print(contragent_list, 'contragent_list')
            # Создадим контрагента:
            contragent = ' '
            if contragent_list:
                contragent = contragent_list.pop(0)
                #contragent_name = 'БНХ РОС'
                if contragent is None:
                    contragent = 'нет данных'
                    dictionaries_models.ContragentsModel.objects.get_or_create(
                        contragent_name = contragent
                        ) 
                if contragent:
                    dictionaries_models.ContragentsModel.objects.get_or_create(
                        contragent_name = contragent
                        )
                #else:
                #    contragent = ' '
                #    dictionaries_models.ContragentsModel.objects.get_or_create(
                #        contragent_name = contragent
                #        )
            else:
                contragent = 'нет данных'
                dictionaries_models.ContragentsModel.objects.get_or_create(
                    contragent_name = contragent
                    )  
            #dictionaries_models.ContragentsModel.objects.get_or_create(
            #    contragent_name = 'БНХ РОС'
            #    )

                    
            for obj in obj_list_el:
            #    # СТАРАЯ ВЕРСИЯ ЗАПИСИ
            #    sale_value = n 
            #    sales_obj = sales_models.Sales.objects.update_or_create(
            #        tyre = obj,
            #        #date_of_sales = date.today(),
            #        date_of_sales = datetime.date(sell_date),
            #        #date_of_sales = datetime.date(2022, 9, 23),
            #        #contragent = dictionaries_models.ContragentsModel.objects.all().last(),
            #        contragent = dictionaries_models.ContragentsModel.objects.get(contragent_name=contragent),
            #        sales_value = int(sale_value),
            #        table = sales_table
            #    )
            #row_value_counter += 1
            #    # END СТАРАЯ ВЕРСИЯ ЗАПИСИ

                sale_value = n 
                obj_list_el_bulk_list.append(sales_models.Sales(
                    tyre = obj,
                    date_of_sales = datetime.date(sell_date),
                    contragent = dictionaries_models.ContragentsModel.objects.get(contragent_name=contragent),
                    sales_value = int(sale_value),
                    table = sales_table
                    )                                
                )
            row_value_counter += 1
        sales_models.Sales.objects.bulk_create(obj_list_el_bulk_list) 


        ####################################################################################################################################################################
        #for n in dictionaries_models.ContragentsModel.objects.all():
        #    print(n.contragent_name)
        #for a in tyretype_row_dict.values(), model_row_dict.values(), params_row_dict.values():
        #    print(a, len(a))
        #for k, v in tyretype_row_dict.items():
        #    print(k, v, ' k, v')
        ## СОЗДАЕМ БАЗОВУЮ ВАЛЮТУ
        currency_chosen_by_hand = dictionaries_models.Currency.objects.get_or_create(
            currency='BYN'
        )
        ####
        ## СОЗДАЕМ СЛОВАРЬ СЕЗОННОСТЬ
        seasons_usage_list = ['летние', 'зимние', 'всесезонные']
        for seas in seasons_usage_list:
            season_created = dictionaries_models.SeasonUsageModel.objects.get_or_create(
                season_usage_name=seas
            )
        ####
        ## СОЗДАЕМ СЛОВАРЬ ОШИПОВКА
        studded_usage_list = ['без шипов', 'с шипами', 'возможность ошиповки']
        for studd in studded_usage_list:
            season_created = dictionaries_models.StuddedUsageModel.objects.get_or_create(
                studded_name=studd
            )
        ####
        #ik = tyres_models.Tyre.objects.all()
        #for n in ik:
        #    print(n.tyre_size.tyre_size, 'TYRE')
        ######################################                      ВАРИАНТ  2  =  СОПОСТАВЛЕНИЕ СОВПАШИХ ШИН С БД
        #### СОПОСТАВЛЕНИЕ СОВПАШИХ СТРОК С С ШИНАМИ В БД:
        ####
        #tyretype_row_dict, model_row_dict, params_row_dict, saless_row_dict, planned_costs_row_dict, semi_variable_costs_row_dict, belarus902price_costs_row_dict, tpsrussiafcaprice_costs_row, tpskazfcaprice_cost_row, tpsmiddleasiafcaprice_costs_row, tpsmiddleasiafcaprice_costs_row 
        #if len(tyretype_row_dict) == len(model_row_dict) and len(tyretype_row_dict) == len(params_row_dict) and len(tyretype_row_dict) == len(saless_row_dict) and len(tyretype_row_dict) == len(planned_costs_row_dict) and len(tyretype_row_dict) == len(semi_variable_costs_row_dict) and len(tyretype_row_dict) == len(belarus902price_costs_row_dict) and len(tyretype_row_dict) == len(tpsrussiafcaprice_costs_row) and len(tyretype_row_dict) == len(tpskazfcaprice_cost_row) and len(tyretype_row_dict) == len(tpsmiddleasiafcaprice_costs_row):
        #if len(tyretype_row_dict) == len(model_row_dict) and len(tyretype_row_dict) == len(params_row_dict) and len(tyretype_row_dict) == len(planned_costs_row_dict) and len(tyretype_row_dict) == len(semi_variable_costs_row_dict) and len(tyretype_row_dict) == len(belarus902price_costs_row_dict) and len(tyretype_row_dict) == len(tpsrussiafcaprice_costs_row) and len(tyretype_row_dict) == len(tpskazfcaprice_cost_row) and len(tyretype_row_dict) == len(tpsmiddleasiafcaprice_costs_row):
        if len(tyretype_row_dict) == len(model_row_dict) and len(tyretype_row_dict) == len(params_row_dict):
            # выбираем именно по строкам, а не просто по порядку. Ведь  ключ -  номер строки:
            for n in tyretype_row_dict.keys():
            #for n in range(0, len(tyretype_row_dict)):
                #print(tyretype_row_dict.get(n), 'tyretype_row_dict.get(n)', len(tyretype_row_dict))
                if tyres_models.Tyre.objects.filter(tyre_model__model=model_row_dict.get(n), tyre_size__tyre_size=tyretype_row_dict.get(n)):
                    #print(model_row_dict.get(n), tyretype_row_dict.get(n), 'tyretype_row_dict.get(n')
                    tyre_mathced_obj = tyres_models.Tyre.objects.filter(tyre_model__model=model_row_dict.get(n), tyre_size__tyre_size=tyretype_row_dict.get(n), tyre_type__tyre_type__in=params_row_dict.get(n))
                    ####for object in tyre_mathced_obj:
                    ####    print(object.tyre_model.model, object.tyre_size.tyre_size,  object.tyre_type.all(), 'n = ', n)
                    for obj in tyre_mathced_obj:              ### берем объекты из кверисета
                        #print(obj.tyre_size.tyre_size, 'tyre_mathced_obj')
                        counter_coinc = 0
                        for param in params_row_dict.get(n):                ### берем параметры перебираем
                            len(params_row_dict.get(n))                     ###  замеряем количество параметров в текущем 
                            dbpar_list = []
                            for p in obj.tyre_type.all():       ###   перебираем параметры в параметрах объекта
                                if param == p.tyre_type:        ### если параметры совпадают
                                    counter_coinc += 1
                                    #matched = param, '=', p.tyre_type
                                dbpar_list.append(p.tyre_type)
                        #print('parsed params = ',  params_row_dict.get(n), 'params in DB=', dbpar_list)
                        if counter_coinc == len(params_row_dict.get(n)):        ###### ЗДЕСЬ ПОЛУЧАЕМ ОБЪЕКТ У КОТОРОГО ВСЕ ПАРАМЕТРЫ СОВПАЛИ СПАРАМЕТРАМИ ШИНЫ В БАЗЕ
                            #print('MATCHED!!', obj.tyre_size.tyre_size)
                            sales_ddict = {}
                            sales_ddict['sales_ddict'] = saless_row_dict.get(n)
                            planned_costs_ddict = {}
                            planned_costs_ddict['planned_costs_ddict'] = planned_costs_row_dict.get(n)
                            semi_variable_ddict = {}
                            semi_variable_ddict['semi_variable_ddict'] = semi_variable_costs_row_dict.get(n)
                            belarus902price_costs_ddict = {}
                            belarus902price_costs_ddict['belarus902price_costs_ddict'] = belarus902price_costs_row_dict.get(n)
                            tpsrussiafcaprice_costs_ddict = {}
                            tpsrussiafcaprice_costs_ddict['tpsrussiafcaprice_costs_ddict'] = tpsrussiafcaprice_costs_row.get(n)
                            tpskazfcaprice_cost_ddict = {}
                            tpskazfcaprice_cost_ddict['tpskazfcaprice_cost_ddict'] = tpskazfcaprice_cost_row.get(n)
                            tpsmiddleasiafcaprice_costs_ddict = {}
                            tpsmiddleasiafcaprice_costs_ddict['tpsmiddleasiafcaprice_costs_ddict'] = tpsmiddleasiafcaprice_costs_row.get(n)
                            indexes_dict = {}
                            indexes_dict['indexes_dict'] = indexes_row_dict.get(n) 
                            season_dict = {}
                            season_dict['season_dict'] = season_row_dict.get(n) 
                            thread_dict = {}
                            thread_dict['thread_dict'] = thread_row_dict.get(n) 
                            ax_dict = {}
                            ax_dict['ax_dict'] = ax_row_dict.get(n) 
                            usability_dict = {}
                            usability_dict['usability_dict'] = usability_row_dict.get(n) 
                            row_parsing_sales_costs_prices_dict[obj] = sales_ddict, planned_costs_ddict, semi_variable_ddict, belarus902price_costs_ddict, tpsrussiafcaprice_costs_ddict, tpskazfcaprice_cost_ddict, tpsmiddleasiafcaprice_costs_ddict, n , indexes_dict, season_dict, thread_dict, ax_dict, usability_dict
            #print('JJJJJJJJJJJJ', current_prices_row, 'current_prices_row')
            #print('LLLLLLLL', indexes_row_dict)
            current_prices_ddict = {}
            for n in range(0, len(current_prices_row)):
                #print(current_prices_row.get(n))
                if tyres_models.Tyre.objects.filter(tyre_model__model=model_row_dict.get(n), tyre_size__tyre_size=tyretype_row_dict.get(n)):
                    tyre_mathced_obj = tyres_models.Tyre.objects.filter(tyre_model__model=model_row_dict.get(n), tyre_size__tyre_size=tyretype_row_dict.get(n), tyre_type__tyre_type__in=params_row_dict.get(n))
                    ####for object in tyre_mathced_obj:
                    ####    print(object.tyre_model.model, object.tyre_size.tyre_size,  object.tyre_type.all(), 'n = ', n)
                    for obj in tyre_mathced_obj:              ### берем объекты из кверисета
                        counter_coinc = 0
                        for param in params_row_dict.get(n):                ### берем параметры перебираем
                            len(params_row_dict.get(n))                     ###  замеряем количество параметров в текущем 
                            dbpar_list = []
                            for p in obj.tyre_type.all():       ###   перебираем параметры в параметрах объекта
                                if param == p.tyre_type:        ### если параметры совпадают
                                    counter_coinc += 1
                                    #matched = param, '=', p.tyre_type
                                dbpar_list.append(p.tyre_type)
                        #print('parsed params = ',  current_prices_row.get(n), 'params in DB=', dbpar_list)
                        
                        if counter_coinc == len(params_row_dict.get(n)):        ###### ЗДЕСЬ ПОЛУЧАЕМ ОБЪЕКТ У КОТОРОГО ВСЕ ПАРАМЕТРЫ СОВПАЛИ СПАРАМЕТРАМИ ШИНЫ В БАЗЕ
                            #print('MATCHED!!', obj)
                            
                            #current_prices_ddict['tpsmiddleasiafcaprice_costs_ddict'] = current_prices_row.get(n)
                            current_prices_ddict[obj] = current_prices_row.get(n)
            #print('current_prices_ddict', current_prices_ddict)
        #for key, value in row_parsing_sales_costs_prices_dict.items():
        #    print(key.tyre_size.tyre_size, 'row_parsing_sales_costs_prices_dict', value)  
        ##############          КЛЮЧЕВАЯ ШТУКА:         !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        #print('row_parsing_sales_costs_prices_dict', row_parsing_sales_costs_prices_dict)           #### !!!!!!!!!!! КОНТРОЛЬНЫЙ СЛОВАРЬ СО СВОДОМ ПОСТРОЧНЫМ СОВПАШИХ С БД ШИН И СПАРСЕННЫМИ ДАННЫМИ как аналог list_of_unique_tyres_objs_cleaned n = номер строки!!!
        #Tyre object (1059) ({'sales_ddict': None}, {'planned_costs_ddict': 27541}, {'semi_variable_ddict': 18935}, {'belarus902price_costs_ddict': 33321}, {'tpsrussiafcaprice_costs_ddict': 37245}, {'tpskazfcaprice_cost_ddict': 37226}, {'tpsmiddleasiafcaprice_costs_ddict': 48877}, 25)
        #'row_parsing_sales_costs_prices_dict':
        # [0]['sales_ddict']                        # объем продаж
        # [1]['planned_costs_ddict']                # себестоимость
        # [2]['semi_variable_ddict']                # плановые затраты
        # [3]['belarus902price_costs_ddict']        # 902,9 РБ
        # [4]['tpsrussiafcaprice_costs_ddict']      # ТПС РФ
        # [5]['tpskazfcaprice_cost_ddict']          # ТПС Казахстан
        # [6]['tpsmiddleasiafcaprice_costs_ddict']  # ТПС Азия
        # [7]['current_prices_ddict']         # действующие цены
        # [8] = номер строки                        # номер строки из спарсингованной таблицы
        # [9]['indexes_dict']                       # ДОПОЛНИТЕЛЬНО индексы  
        # [10]['season_dict']                        # ДОПОЛНИТЕЛЬНО сезонность 
        # [11]['thread_dict']                       # ДОПОЛНИТЕЛЬНО рисунок протекора
        # [12]['ax_dict']                           # ДОПОЛНИТЕЛЬНО ось   
        # [13]['usability_dict']                    # ДОПОЛНИТЕЛЬНО применяемость        
#{'sales_ddict': None},                             0
#{'planned_costs_ddict': None},                     1
#{'semi_variable_ddict': 19311},                    2
#{'belarus902price_costs_ddict': 34589},            3
#{'tpsrussiafcaprice_costs_ddict': None},           4
#{'tpskazfcaprice_cost_ddict': None},               5
#{'tpsmiddleasiafcaprice_costs_ddict': None},       6
#25,                                                7
#{'indexes_dict': '164J'},                          8
#{'season_dict': None},                             9
#{'thread_dict': 'Повышенная проходимость'},        10
#{'ax_dict': None},                                 11
#{'usability_dict': 'самосвал'})                    12 
        #########
        #'current_prices_ddict'                     # действующие цены - отдельный словарь
        #for k, v in row_parsing_sales_costs_prices_dict.items():       # ПРЕДПРОСМОТР СПИСКА СОВПАВШИХ ШИН И ДАННЫХ
        #    parrams = []
        #    for par in k.tyre_type.all():
        #        parrams.append(par.tyre_type)
        #    print(k.tyre_size.tyre_size, k.tyre_model.model, parrams, v)
        ### глобальная дата документа
        if date_period_of_doc:
            date_period_table = date_period_of_doc
        else:
            date_period_table = datetime.now()
        
        #print('date_period_table TTT', date_period_table)
        # ЗАБРАСЫВАЕМ ПОЛНЫЕ ЗАТРАТЫ:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            #print(obj_list_el[1]['planned_costs_ddict'] , 'KKK')
            if obj_list_el[1]['planned_costs_ddict']:
                    planned_costs_object = prices_models.PlannedCosstModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el[1]['planned_costs_ddict'],
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        # ЗАБРАСЫВАЕМ ПЛАНОВУЮ СЕБЕСТОИМОСТЬ:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            if obj_list_el[2]['semi_variable_ddict']:
                    semi_variable_costs_object = prices_models.SemiVariableCosstModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el[2]['semi_variable_ddict'],
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        # ЗАБРАСЫВАЕМ прейскуранты №№9, 902:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            if obj_list_el[3]['belarus902price_costs_ddict']:
                    belarus902price_costs_object = prices_models.Belarus902PriceModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el[3]['belarus902price_costs_ddict'],
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        # ЗАБРАСЫВАЕМ даннын по  ТПС РФ FCA:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            if obj_list_el[4]['tpsrussiafcaprice_costs_ddict']:
                    tpsrussiafcaprice_object = prices_models.TPSRussiaFCAModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el[4]['tpsrussiafcaprice_costs_ddict'],
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        
        # ЗАБРАСЫВАЕМ даннын по ТПС Казахстан FCA:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            if obj_list_el[5]['tpskazfcaprice_cost_ddict']:
                    tpskazfcaprice_object = prices_models.TPSKazFCAModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el[5]['tpskazfcaprice_cost_ddict'],
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        # ЗАБРАСЫВАЕМ даннын по ТПС Средняя Азия, Закавказье, Молдова FCA:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            if obj_list_el[6]['tpsmiddleasiafcaprice_costs_ddict']:
                    tpsmiddleasiafcaprice_object = prices_models.TPSMiddleAsiaFCAModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el[6]['tpsmiddleasiafcaprice_costs_ddict'],
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        #print(row_parsing_sales_costs_prices_dict.items(), 'row_parsing_sales_costs_prices_dict.items()')
        # ЗАБРАСЫВАЕМ даннын по Действующие цены:
        #for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
        for key, obj_list_el in current_prices_ddict.items():
            #print('row_parsing_sales_costs_prices_dict', key, obj_list_el )             #current_prices_ddict['tpsmiddleasiafcaprice_costs_ddict']
            if obj_list_el:
                    currentpricesprice_object = prices_models.CurrentPricesModel.objects.update_or_create(
                        tyre = key,     
                        price = obj_list_el,
                        #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ
                        date_period = date_period_table,   
                        currency = currency_chosen_by_hand[0]         
                    )
        ## СОЗДАЕМ объектыTyreAddedFeatureModel:
        ## ЗАБРАСЫВАЕМ ИНДЕКСЫ, сезонность, рисунок протектора, ось, применяемость ДОПОЛНИТЕЛЬНОЕ в модель TyreAddedFeatureModel - дополнительная таблица к таблице модели Tyres:
        #if obj_list_el[9] or obj_list_el[10] or obj_list_el[11] or obj_list_el[12] or obj_list_el[13]:
        for key, obj_list_el in row_parsing_sales_costs_prices_dict.items():
            #print('KEY', key)
            #if obj_list_el[8]['indexes_dict'] and obj_list_el[9]['season_dict'] and obj_list_el[10]['thread_dict'] and obj_list_el[11]['ax_dict'] and obj_list_el[12]['usability_dict']:
            season_usage = dictionaries_models.SeasonUsageModel.objects.filter(season_usage_name=obj_list_el[9]['season_dict']) 
            if season_usage:
                season_usage = season_usage[0]
            else:
                season_usage = None 
            #studded_usage = dictionaries_models.StuddedUsageModel.objects.filter(season_usage_name=obj_list_el[9]['ЗДЕСЬ ДОЛЖЕН БЫТЬ СПАРСЕННЫЙ СЛОВАРЬ С ОШИПОВКОЙ]) 
            #if studded_usage:
            #    studded_usage = studded_usagee[0]
            #else:
            #    studded_usage = None 
            #print('key ==', key, 'indexes_list===', obj_list_el[8]['indexes_dict'], 'season_usage===', season_usage, 'tyre_thread===', obj_list_el[10]['thread_dict'], 'ax===', obj_list_el[11]['ax_dict'] , 'usability===', obj_list_el[12]['usability_dict'])
            try:
                tyre_added_feature_object = tyres_models.TyreAddedFeatureModel.objects.update_or_create(
                    tyre = key,   
                    indexes_list = obj_list_el[8]['indexes_dict'],
                    season_usage = season_usage,
                    tyre_thread = obj_list_el[10]['thread_dict'],
                    ax = obj_list_el[11]['ax_dict'],
                    usability = obj_list_el[12]['usability_dict'],
                    #studded_usage = studded_usage          
                    #date_period = date_from_table,                 # ЗДЕСЬ ПРОПИСЫВАТЬ ДАТУ ВМЕСТО ЗАГЛУШКИ                        
                )
            except:
                pass
    ## соберем Tyre_Sale объекты:
    #for obj in tyres_models.Tyre.objects.all():
    #    tyr_sal = sales_models.Tyre_Sale.objects.update_or_create(
    #        tyre = obj,
    #        table = sales_table
    #    )
        for obj in tyres_models.Tyre.objects.all():
            tyr_sal = sales_models.Tyre_Sale.objects.update_or_create(
                tyre = obj,
                table = sales_table
            )
    # Создаем справочники по базовым валютам:
    base_currencies = ['RUB', 'BYN', 'USD', 'EURO']
    for curr in base_currencies:
        dictionaries_models.Currency.objects.get_or_create(
            currency = curr
        )
    ###
    #### соберем объекты ABC из шин и объектов Sales:
    #0) создадим объект таблицы AbcxyzTable для проверки:
    #table_id = self.request.session.get('table_id')             
    #abc_table_get = abc_table_xyz_models.AbcxyzTable.objects.filter(pk=table_id)
    #if abc_table_get:
    #    abc_table = abc_table_get[0]
    #else:
    ##    print('iiii', abc_table_xyz_models.AbcxyzTable.objects.update_or_create())
    #    abc_table_queryset = abc_table_xyz_models.AbcxyzTable.objects.update_or_create()
    #    abc_table = abc_table_queryset[0]
    # 1) возмем все объекты шин:
    for obj in tyres_models.Tyre.objects.all():
        tyre_obj = obj
    # 2) возмем все объекты продаж:
        sales_obj_set = sales_models.Sales.objects.filter(tyre=tyre_obj)
        #print(sales__obj_set)       # <QuerySet [<Sales: Sales object (999)>, <Sales: Sales object (1017)>]>
        #print(list(sales__obj_set))
    ## 3) создадим объекты модели Abcxyz
    #    abc_obj = abc_table_xyz_models.Abcxyz.objects.filter(table=abc_table, tyre=tyre_obj)
    #    abc_obj_set = abc_table_xyz_models.Abcxyz.objects.update_or_create(
    #        table=abc_table,
    #        tyre=tyre_obj,
    #    )
    #    for sales_obj in sales_obj_set:
    #        abc_obj_set[0].sales.add(sales_obj)
        #print(abc_obj_set[0].sales.all(), 'JJJJJJJJ')
    #####if abc_table:
    #####    #print(abc_table, '==', abc_table.tyre_total, abc_table.list_of_total_sales_of_of_tyre_in_period())
    #####    pass
    #for obj in abc_table_xyz_models.Abcxyz.objects.all():
    #    #print(obj.total_sales_of_of_tyre_in_period(), obj.percent_in_total_amount())
    #    print(obj.abc_group[0])
    #0) создадим объект таблицы ComparativeAnalysisTableModel для проверки:
    # старый вариант:
    #table_id = self.request.session.get('table_id')             
    #comparative_analysis_table_get = prices_models.ComparativeAnalysisTableModel.objects.filter(pk=table_id)
    #if comparative_analysis_table_get:
    #    comparative_analysis_table = comparative_analysis_table_get[0]
    #else:
    #    comparative_analysis_table_queryset = prices_models.ComparativeAnalysisTableModel.objects.update_or_create()
    #    comparative_analysis_table = comparative_analysis_table_queryset[0]
    # новый вариант:
    list_of_market_price_names = ['belarus', 'russia']
    for market_name in list_of_market_price_names:
        prices_models.ComparativeAnalysisTableModel.objects.get_or_create(market_table=market_name)
    #for key, valluuee in row_parsing_sales_costs_prices_dict.items():
    #    print(key.tyre_size.tyre_size, key.tyre_model.model, valluuee[8], valluuee[9], valluuee[10], valluuee[12])
    # 1) возмем все объекты шин:
    #for key in row_parsing_sales_costs_prices_dict.keys():
    #    print('key', key, key.tyre_model.model, key.tyre_size.tyre_size, key.tyre_group.all(), key.tyre_type.all())
    for key in row_parsing_sales_costs_prices_dict.keys():
        tyre_obj = key
    #    print('tyre_obj', tyre_obj.tyre_size.tyre_size)
    # 2.1) возмем все объекты плановой себестоимости:
        planned_costs_obj_set = prices_models.PlannedCosstModel.objects.filter(tyre=tyre_obj)            # !!!! ФИЛЬТР ВСЕХ ОБЪЕКТОВ + ДОБАВИТЬ ФИЛЬТР ПО ПЕРИОДУ   ===== date_period
        #planned_costs_obj_set = prices_models.PlannedCosstModel.objects.get(tyre=tyre_obj)                  # ОДИН объект на дату + ДОБАВИТЬ ФИЛЬТР ПО ПЕРИОДУ   ===== date_period
        if planned_costs_obj_set:
            planned_costs_obj_set = planned_costs_obj_set[0]
        else:       # если пустой кверисет
            planned_costs_obj_set = None
    # 2.2) возмем все объекты прямые затраты:
        semi_variable_costs_obj_set = prices_models.SemiVariableCosstModel.objects.filter(tyre=tyre_obj)
        #semi_variable_costs_obj_set = prices_models.SemiVariableCosstModel.objects.get(tyre=tyre_obj)
        if semi_variable_costs_obj_set:
            semi_variable_costs_obj_set = semi_variable_costs_obj_set[0]
        else:       # если пустой кверисет
            semi_variable_costs_obj_set = None
    # 2.3) возмем все объекты прейскуранты №№9, 902:
        belarus902price_obj_set = prices_models.Belarus902PriceModel.objects.filter(tyre=tyre_obj)
        #belarus902price_obj_set = prices_models.Belarus902PriceModel.objects.get(tyre=tyre_obj)
        if belarus902price_obj_set:
            belarus902price_obj_set = belarus902price_obj_set[0]
            #print('belarus902price_obj_set', belarus902price_obj_set.price)
        else:       # если пустой кверисет
            belarus902price_obj_set = None
    # 2.4) возмем все объекты прейскуранты ТПС РФ FCA:
        tpsrussiafcaprice_obj_set = prices_models.TPSRussiaFCAModel.objects.filter(tyre=tyre_obj)
        #tpsrussiafcaprice_obj_set = prices_models.TPSRussiaFCAModel.objects.get(tyre=tyre_obj)
        if  tpsrussiafcaprice_obj_set:
            tpsrussiafcaprice_obj_set = tpsrussiafcaprice_obj_set[0]
        else:       # если пустой кверисет
            tpsrussiafcaprice_obj_set = None
    # 2.5) возмем все объекты прейскуранты ТПС Казахстан FCA:
        tpskazfcaprice_obj_set = prices_models.TPSKazFCAModel.objects.filter(tyre=tyre_obj)
        #tpskazfcaprice_obj_set = prices_models.TPSKazFCAModel.objects.get(tyre=tyre_obj)
        if  tpskazfcaprice_obj_set:
            tpskazfcaprice_obj_set = tpskazfcaprice_obj_set[0]
        else:       # если пустой кверисет
            tpskazfcaprice_obj_set = None
    # 2.6) возмем все объекты прейскуранты ТПС Средняя Азия, Закавказье, Молдова FCA:
        tpsmiddleasiafcaprice_obj_set = prices_models.TPSMiddleAsiaFCAModel.objects.filter(tyre=tyre_obj)
        #tpsmiddleasiafcaprice_obj_set = prices_models.TPSMiddleAsiaFCAModel.objects.get(tyre=tyre_obj)
        if tpsmiddleasiafcaprice_obj_set:
            tpsmiddleasiafcaprice_obj_set = tpsmiddleasiafcaprice_obj_set[0]
        else:       # если пустой кверисет
            tpsmiddleasiafcaprice_obj_set = None
    # 2.7) возмем все объекты Действующие цены:
        currentpricesprice_obj_set = prices_models.CurrentPricesModel.objects.filter(tyre=tyre_obj)
        #currentpricesprice_obj_set = prices_models.CurrentPricesModel.objects.get(tyre=tyre_obj)
        if currentpricesprice_obj_set:
            currentpricesprice_obj_set = currentpricesprice_obj_set[0]
        else:       # если пустой кверисет
            currentpricesprice_obj_set = None
        #print(planned_costs_obj_set.price, semi_variable_costs_obj_set.price, belarus902price_obj_set.price, 'HHH')
    # 3) создадим объекты модели ComparativeAnalysisTyresModel
        #defaults = dict(
        #    planned_costs=planned_costs.price,
        #    semi_variable_prices=semi_variable_prices,
        #    belarus902price=belarus902price,
        #    tpsrussiafcaprice=0,
        #    tpskazfcaprice=0,
        #    tpsmiddleasiafcaprice=0,
        #    currentpricesprice=0,                
        #)
        ## новое дополнение:
        #for comparative_analysis_table in prices_models.ComparativeAnalysisTableModel.objects.all():
        #    comparative_analysis_tyres_obj_set = prices_models.ComparativeAnalysisTyresModel.objects.update_or_create(
        #        table=comparative_analysis_table,
        #        tyre=tyre_obj,
        #        planned_costs=planned_costs_obj_set,
        #        semi_variable_prices=semi_variable_costs_obj_set,
        #        belarus902price=belarus902price_obj_set,
        #        tpsrussiafcaprice=tpsrussiafcaprice_obj_set,
        #        tpskazfcaprice=tpskazfcaprice_obj_set,
        #        tpsmiddleasiafcaprice=tpsmiddleasiafcaprice_obj_set,
        #        currentpricesprice=currentpricesprice_obj_set,
        #        #defaults={'semi_variable_prices': 0, 'belarus902price': 0, 'tpsrussiafcaprice': 0, 'tpskazfcaprice': 0, 'tpsmiddleasiafcaprice': 0, 'currentpricesprice': 0, 'currentpricesprice': 0},
        #        )
        #print('1', tyre_obj, '2', planned_costs_obj_set, '3', semi_variable_costs_obj_set, '4', belarus902price_obj_set, '5', tpsrussiafcaprice_obj_set, '6', tpsmiddleasiafcaprice_obj_set, '7', tpsmiddleasiafcaprice_obj_set, '8', currentpricesprice_obj_set  )
        comparative_analysis_tyres_obj_set = prices_models.ComparativeAnalysisTyresModel.objects.update_or_create(
            #table=comparative_analysis_table,
            tyre=tyre_obj,
            planned_costs=planned_costs_obj_set,
            semi_variable_prices=semi_variable_costs_obj_set,
            belarus902price=belarus902price_obj_set,
            tpsrussiafcaprice=tpsrussiafcaprice_obj_set,
            tpskazfcaprice=tpskazfcaprice_obj_set,
            tpsmiddleasiafcaprice=tpsmiddleasiafcaprice_obj_set,
            currentpricesprice=currentpricesprice_obj_set,
            #defaults={'semi_variable_prices': 0, 'belarus902price': 0, 'tpsrussiafcaprice': 0, 'tpskazfcaprice': 0, 'tpsmiddleasiafcaprice': 0, 'currentpricesprice': 0, 'currentpricesprice': 0},
            sale_data = date_period_table
            )
        for comparative_analysis_table in prices_models.ComparativeAnalysisTableModel.objects.all():
            comparative_analysis_tyres_obj_set[0].table.add(comparative_analysis_table)
    ############################################ Запись данных в существующий файл в столбец:
    from openpyxl import load_workbook
    excel_file = load_workbook('Tyres.xlsx')
    excel_sheet = excel_file["Holidays 2019"]
    #print(excel_file.sheetnames, 'TTTTTTTTTTTTTTTTTT')
    ###### получить количество объектов Sales:
    row_value = 0
    for ob_val in sales_models.Sales.objects.all():
        row_value += 1
    
    #генератор для хождения по строкам:    
    def raw_generator(n, stop):
        while True:
            if n > stop:
                raise StopIteration
            yield n
            n += 1
    max_raw = row_value     
    i = 2       # со второй строки        
    row_curr = raw_generator(i, max_raw+i)
    for sales_obj in sales_models.Sales.objects.all():
        val = next(row_curr)
        excel_sheet['E1'] = 'Tyre Size'
        excel_sheet.cell(row=val, column=5).value = sales_obj.tyre.tyre_size.tyre_size
        excel_sheet['F1'] = 'Model'
        excel_sheet.cell(row=val, column=6).value = sales_obj.tyre.tyre_model.model
        excel_sheet['G1'] = 'Param'
        str_obj_param_list = []
        str_obj_param_str = ''
        for type_obj in sales_obj.tyre.tyre_type.all():
            str_obj_param_list.append(type_obj.tyre_type)
        str_obj_param_str = ', '.join(str_obj_param_list)
        #print(str_obj_param_str)
        excel_sheet.cell(row=val, column=7).value = str_obj_param_str
        excel_sheet['H1'] = 'Sales value'
        excel_sheet.cell(row=val, column=8).value = sales_obj.sales_value
        excel_sheet['I1'] = 'Date_of_sales'
        excel_sheet.cell(row=val, column=9).value = sales_obj.date_of_sales
        excel_sheet['J1'] = 'Contragent'
        excel_sheet.cell(row=val, column=10).value = sales_obj.contragent.contragent_name
        excel_file.save(filename="Tyres.xlsx") 


    try:
        get_saved_file_form_a.close()
    except:
        pass
    try:
        get_saved_file_form_b.close()
    except:
        pass
    
    ## если не химкурьер импорт - то вернуть флаг что был импорт не химкурьер BИНАЧЕ- смотри выше return
    return copy_file_created,  list_of_sheet_potential_names_var_list_is 







def main(self):
    read_from_file()

if __name__ == "__main__":
    main()    